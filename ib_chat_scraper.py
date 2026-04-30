"""
IB Chat Scraper — Continuous daemon
====================================

Connects to a logged-in Bloomberg Terminal via blpapi and subscribes to all IB
chat rooms the user is a member of, then filters down to the counterparties
you care about. Appends every kept message to:

  1. ib_chats.jsonl  — append-only log, one JSON object per line (primary store)
  2. ib_chats.xlsx   — human-readable Excel, rebuilt from JSONL on each run

Filtering:
  Configure scrape_filters.json (or pass --config <path>) to whitelist by:
    - uuids        : list of Bloomberg UUIDs (most stable identifier)
    - sender_names : list of name substrings (e.g. ["JOHN SMITH", "JANE DOE"])
    - firms        : list of firm-name substrings (e.g. ["GOLDMAN", "MORGAN STANLEY"])
    - rooms        : list of room-name substrings (e.g. ["GS FX OPTS", "MS FX EM"])
  
  A message is kept if ANY filter category matches. If no filters are configured,
  ALL messages are kept (legacy behavior).
  
  Outbound messages from you are always captured if they appear in a room/with a
  counterparty that matches any filter — so your replies stay in context.

Design goals:
  - Run continuously, survive Bloomberg disconnects with auto-reconnect
  - Cumulative storage: each run adds new messages to the existing log
  - Deduplicate by msgId so reconnects don't double-write
  - Crash-safe: JSONL append is atomic per-line, Excel rebuilds from JSONL
  - Minimal interpretation: capture raw message + metadata, leave parsing for later

Usage:
  python ib_chat_scraper.py                    # run continuously with config file
  python ib_chat_scraper.py --once             # one-shot snapshot then exit
  python ib_chat_scraper.py --rebuild-xlsx     # rebuild Excel from JSONL only
  python ib_chat_scraper.py --config my.json   # use custom filter config
  python ib_chat_scraper.py --no-filter        # capture everything (legacy mode)
  python ib_chat_scraper.py --discover         # log all senders/rooms seen, no filtering

Output files (created in script directory):
  ib_chats.jsonl       cumulative message log
  ib_chats.xlsx        formatted spreadsheet for review
  ib_scraper.log       runtime log
  scrape_filters.json  filter configuration (auto-created on first run with example)
  ib_seen_senders.json discovery log of every sender/room seen (for tuning filters)
"""

import argparse
import json
import os
import re
import signal
import sys
import threading
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

# Constants
HERE = Path(os.path.dirname(os.path.abspath(__file__)))
JSONL_PATH = HERE / 'ib_chats.jsonl'
XLSX_PATH = HERE / 'ib_chats.xlsx'
LOG_PATH = HERE / 'ib_scraper.log'
FILTER_PATH = HERE / 'scrape_filters.json'
SEEN_PATH = HERE / 'ib_seen_senders.json'

# Excel rebuild interval (seconds) — don't rebuild on every message, batch up
XLSX_REBUILD_INTERVAL = 60

# Bloomberg connection
BBG_HOST = 'localhost'
BBG_PORT = 8194


# ============================================================
# FILTER CONFIG
# ============================================================

EXAMPLE_FILTER = {
    "_comment": "Whitelist filter — a message is kept if it matches ANY category. Leave a category empty to skip that check. To capture EVERYTHING, set capture_all=true. Use --discover for a few minutes first to find UUIDs/names/firms/rooms in your environment.",
    "capture_all": False,
    "uuids": [
        # 12345678,
        # 87654321,
    ],
    "sender_names": [
        # "JOHN SMITH",
        # "JANE DOE"
    ],
    "firms": [
        # "GOLDMAN SACHS",
        # "MORGAN STANLEY",
        # "JP MORGAN"
    ],
    "rooms": [
        # "GS FX OPTS NY",
        # "MS FX EM"
    ],
    "always_keep_outbound_in_matched_rooms": True
}


class ChatFilter:
    """Whitelist filter for IB messages."""

    def __init__(self, config: dict):
        self.capture_all = bool(config.get('capture_all', False))
        # Normalize all string filters to upper for case-insensitive matching
        self.uuids = set(int(u) for u in config.get('uuids', []) if str(u).strip())
        self.sender_names = [s.upper().strip() for s in config.get('sender_names', []) if s.strip()]
        self.firms = [s.upper().strip() for s in config.get('firms', []) if s.strip()]
        self.rooms = [s.upper().strip() for s in config.get('rooms', []) if s.strip()]
        self.keep_outbound_in_matched = bool(
            config.get('always_keep_outbound_in_matched_rooms', True)
        )
        # Cache: rooms we've already decided contain a matched counterparty
        # (keyed by roomName, value=True/False). Lets us keep your replies.
        self._matched_rooms = set()

    def has_any_filter(self) -> bool:
        return bool(self.uuids or self.sender_names or self.firms or self.rooms)

    def describe(self) -> str:
        if self.capture_all:
            return 'capture_all=True (no filtering)'
        if not self.has_any_filter():
            return 'no filters configured -> capturing everything'
        parts = []
        if self.uuids:        parts.append(f'{len(self.uuids)} UUIDs')
        if self.sender_names: parts.append(f'{len(self.sender_names)} sender names')
        if self.firms:        parts.append(f'{len(self.firms)} firms')
        if self.rooms:        parts.append(f'{len(self.rooms)} rooms')
        return ' + '.join(parts)

    def should_keep(self, rec: dict) -> bool:
        """Decide whether to persist this message based on the configured filters."""
        if self.capture_all:
            return True
        if not self.has_any_filter():
            return True  # no filter = keep all (back-compat)

        room = (rec.get('roomName') or '').upper()
        sender = (rec.get('senderName') or '').upper()
        firm = (rec.get('firmName') or '').upper()
        is_outbound = bool(rec.get('isOutbound'))

        # Try to coerce UUID to int (may arrive as string)
        try:
            uuid = int(rec.get('senderUuid') or 0)
        except (ValueError, TypeError):
            uuid = 0

        matched = False

        # UUID match (exact)
        if self.uuids and uuid in self.uuids:
            matched = True

        # Sender name match (substring, case-insensitive)
        if not matched and self.sender_names:
            if any(sn in sender for sn in self.sender_names):
                matched = True

        # Firm match (substring, case-insensitive)
        if not matched and self.firms:
            if any(fm in firm for fm in self.firms):
                matched = True

        # Room match (substring, case-insensitive)
        if not matched and self.rooms:
            if any(rm in room for rm in self.rooms):
                matched = True

        if matched:
            # Remember this room contains a matched party — keep your replies
            if room:
                self._matched_rooms.add(room)
            return True

        # Outbound message in a previously-matched room?
        if is_outbound and self.keep_outbound_in_matched and room in self._matched_rooms:
            return True

        return False


def load_filter_config(path: Path = None) -> dict:
    """Load filter config from JSON. Auto-create example if missing."""
    path = path or FILTER_PATH
    if not path.exists():
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(EXAMPLE_FILTER, f, indent=2)
        log(f'Created example filter config at {path.name}')
        log(f'  Edit it to specify which counterparties to capture, then re-run.')
        return EXAMPLE_FILTER
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ============================================================
# LOGGING
# ============================================================

def log(msg, *, also_print=True):
    """Append to log file, optionally print."""
    ts = datetime.now().strftime('%H:%M:%S')
    line = f'[{ts}] {msg}'
    try:
        with open(LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception:
        pass
    if also_print:
        print(line, flush=True)


# ============================================================
# JSONL STORAGE — append-only log
# ============================================================

def load_seen_ids():
    """Read JSONL once at startup to populate dedup set of msgIds we've already written."""
    seen = set()
    if not JSONL_PATH.exists():
        return seen
    try:
        with open(JSONL_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    if 'msgId' in rec:
                        seen.add(rec['msgId'])
                except json.JSONDecodeError:
                    continue
    except Exception as e:
        log(f'WARN: load_seen_ids failed: {e}')
    log(f'Loaded {len(seen)} previously-seen message IDs from {JSONL_PATH.name}')
    return seen


def append_jsonl(record):
    """Append one record to JSONL file. Atomic per-line."""
    try:
        with open(JSONL_PATH, 'a', encoding='utf-8') as f:
            f.write(json.dumps(record, ensure_ascii=False) + '\n')
        return True
    except Exception as e:
        log(f'ERROR: append_jsonl failed: {e}')
        return False


def load_all_records():
    """Read entire JSONL into a list of dicts. Used for Excel rebuild."""
    records = []
    if not JSONL_PATH.exists():
        return records
    with open(JSONL_PATH, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return records


# ============================================================
# EXCEL REBUILD
# ============================================================

def rebuild_xlsx():
    """Rebuild Excel file from current JSONL contents. Called periodically."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log('WARN: openpyxl not installed, skipping xlsx rebuild. pip install openpyxl')
        return

    records = load_all_records()
    if not records:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'IB Chats'

    headers = ['Timestamp (UTC)', 'Date', 'Time', 'Room', 'Sender', 'Firm',
               'Direction', 'Message', 'msgId']
    ws.append(headers)

    # Header styling
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    # Sort records by timestamp (oldest first, newest at bottom — append-style)
    records.sort(key=lambda r: r.get('timestamp', ''))

    out_fill = PatternFill('solid', fgColor='FFF2CC')   # outbound = yellow
    in_fill = PatternFill('solid', fgColor='FFFFFF')    # inbound = white

    for rec in records:
        ts_str = rec.get('timestamp', '')
        try:
            ts = datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
            date_str = ts.strftime('%Y-%m-%d')
            time_str = ts.strftime('%H:%M:%S')
        except Exception:
            date_str = ''
            time_str = ''

        direction = 'OUT' if rec.get('isOutbound') else 'IN'
        row = [
            ts_str,
            date_str,
            time_str,
            rec.get('roomName', ''),
            rec.get('senderName', ''),
            rec.get('firmName', ''),
            direction,
            rec.get('body', ''),
            rec.get('msgId', ''),
        ]
        ws.append(row)
        # Color rows by direction
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).fill = (
                out_fill if rec.get('isOutbound') else in_fill
            )

    # Column widths
    widths = [22, 12, 10, 28, 22, 24, 10, 80, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Wrap message column
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Save atomically: write to .tmp then rename
    tmp = XLSX_PATH.with_suffix('.xlsx.tmp')
    wb.save(tmp)
    tmp.replace(XLSX_PATH)
    log(f'Excel rebuilt: {len(records)} messages -> {XLSX_PATH.name}')


# ============================================================
# BLOOMBERG IB SCRAPER
# ============================================================

class IBScraper:
    """Subscribes to IB chat events from a logged-in Bloomberg Terminal."""

    # Bloomberg services to try, in order
    # //blp/msgscrape is the documented subscription service for IB chats
    # Field names are based on standard Bloomberg conventions; may need adjustment
    SERVICE_NAME = '//blp/msgscrape'

    def __init__(self, chat_filter: ChatFilter, discover_mode: bool = False):
        self.filter = chat_filter
        self.discover = discover_mode
        self.seen_ids = load_seen_ids()
        self._running = False
        self._session = None
        self._lock = threading.Lock()
        self._last_xlsx_rebuild = 0
        self._new_msgs_since_rebuild = 0
        # Discovery tracking: count messages per (sender, firm, room, uuid)
        self._discovery = defaultdict(lambda: {
            'sender': '', 'firm': '', 'room': '', 'uuid': 0,
            'inbound_msgs': 0, 'outbound_msgs': 0, 'first_seen': '', 'last_seen': ''
        })

    def _connect(self):
        """Establish blpapi session. Returns True on success."""
        try:
            import blpapi
        except ImportError:
            log('FATAL: blpapi not installed. pip install blpapi')
            log('  See https://www.bloomberg.com/professional/support/api-library/ for installer')
            return False

        opts = blpapi.SessionOptions()
        opts.setServerHost(BBG_HOST)
        opts.setServerPort(BBG_PORT)
        opts.setAutoRestartOnDisconnection(True)
        opts.setNumStartAttempts(3)

        self._session = blpapi.Session(opts)
        if not self._session.start():
            log(f'ERROR: cannot connect to Bloomberg at {BBG_HOST}:{BBG_PORT}')
            log('  Make sure Bloomberg Terminal is running and logged in')
            return False
        log(f'Connected to Bloomberg at {BBG_HOST}:{BBG_PORT}')
        return True

    def _open_msgscrape(self):
        """Open the IB message service. Different Bloomberg builds use different names."""
        candidates = [
            '//blp/msgscrape',      # primary documented name
            '//blp/msgscrape-beta', # some installs
            '//blp/ibmsgscrape',    # alternative
        ]
        for svc in candidates:
            try:
                if self._session.openService(svc):
                    log(f'Opened service: {svc}')
                    self.SERVICE_NAME = svc
                    return True
            except Exception as e:
                log(f'  service {svc} failed: {e}')
        log('ERROR: could not open any IB message service. Your Terminal may not have IB API enabled.')
        log('  Contact your Bloomberg rep to confirm //blp/msgscrape is licensed for your account.')
        return False

    def _subscribe(self):
        """Subscribe to all IB chat traffic for this user."""
        import blpapi
        try:
            subs = blpapi.SubscriptionList()
            # Subscribe to all messages — Bloomberg routes them by your UUID automatically
            # Topic format varies by service; "//blp/msgscrape/all" is the common pattern
            topic = f'{self.SERVICE_NAME}/all'
            subs.add(topic, '', '', blpapi.CorrelationId('IB_ALL'))
            self._session.subscribe(subs)
            log(f'Subscribed to: {topic}')
            return True
        except Exception as e:
            log(f'ERROR: subscribe failed: {e}')
            return False

    def _extract_record(self, msg):
        """Pull fields out of a Bloomberg message object into our record format.
        
        Field names below are standard Bloomberg IB conventions but may vary by 
        Terminal build. If a field is missing, log it and we'll iterate.
        """
        rec = {
            'capturedAt': datetime.now(timezone.utc).isoformat(),
        }
        # Try standard field names — log any we can't find
        field_map = {
            'msgId':       ['MSG_ID', 'MESSAGE_ID', 'msgId'],
            'timestamp':   ['MSG_TIMESTAMP', 'TIMESTAMP_NANOS', 'EVENT_TIME'],
            'roomName':    ['ROOM_NAME', 'CHAT_ROOM', 'ROOM'],
            'roomId':      ['ROOM_ID', 'CHAT_ROOM_ID'],
            'senderUuid':  ['SENDER_UUID', 'FROM_UUID', 'SENDER_ID'],
            'senderName':  ['SENDER_NAME', 'FROM_NAME', 'SENDER'],
            'firmName':    ['SENDER_FIRM', 'FIRM_NAME', 'FROM_FIRM'],
            'body':        ['MSG_BODY', 'BODY', 'MESSAGE_TEXT', 'MSG_TEXT'],
            'isOutbound':  ['IS_OUTBOUND', 'OUTBOUND', 'IS_SELF'],
        }
        unfound = []
        for our_name, candidates in field_map.items():
            value = None
            for cand in candidates:
                if msg.hasElement(cand):
                    try:
                        elem = msg.getElement(cand)
                        if elem.datatype() == 0:  # null
                            continue
                        # Try string first, fall back to other types
                        try:
                            value = elem.getValueAsString()
                        except Exception:
                            try:
                                value = elem.getValue()
                            except Exception:
                                value = str(elem)
                        break
                    except Exception:
                        continue
            if value is None:
                unfound.append(our_name)
                value = ''
            rec[our_name] = value

        if unfound:
            log(f'  WARN: missing fields {unfound} — message logged with empty values')
            log(f'  raw msg dump (first 500 chars): {str(msg)[:500]}')

        # Generate synthetic msgId if Bloomberg didn't provide one
        if not rec.get('msgId'):
            body_excerpt = (rec.get('body') or '')[:50]
            rec['msgId'] = f"synth-{rec.get('timestamp', '')}-{hash(body_excerpt) & 0xFFFFFFFF:08x}"

        return rec

    def _track_discovery(self, rec: dict):
        """Track sender/firm/room combos seen — for tuning filters later."""
        sender = (rec.get('senderName') or '').strip()
        firm = (rec.get('firmName') or '').strip()
        room = (rec.get('roomName') or '').strip()
        try:
            uuid = int(rec.get('senderUuid') or 0)
        except (ValueError, TypeError):
            uuid = 0
        # Key by uuid+room so each rep-in-room combo is its own entry
        key = f'{uuid}|{room}'
        ts = rec.get('timestamp', '')
        info = self._discovery[key]
        info['sender'] = sender or info['sender']
        info['firm'] = firm or info['firm']
        info['room'] = room or info['room']
        info['uuid'] = uuid or info['uuid']
        if rec.get('isOutbound'):
            info['outbound_msgs'] += 1
        else:
            info['inbound_msgs'] += 1
        if not info['first_seen']:
            info['first_seen'] = ts
        info['last_seen'] = ts

    def _save_discovery(self):
        """Persist the discovery log so user can build filters from it."""
        try:
            entries = sorted(
                self._discovery.values(),
                key=lambda d: -(d['inbound_msgs'] + d['outbound_msgs'])
            )
            with open(SEEN_PATH, 'w', encoding='utf-8') as f:
                json.dump(entries, f, indent=2)
        except Exception as e:
            log(f'  WARN: save_discovery failed: {e}')

    def _process_message(self, msg):
        """Handle one IB message: dedup, filter, persist."""
        rec = self._extract_record(msg)

        msg_id = rec.get('msgId')
        if not msg_id:
            log('  skipping message with no msgId')
            return

        with self._lock:
            if msg_id in self.seen_ids:
                return  # already captured

            # Always track discovery — even messages we filter out
            self._track_discovery(rec)

            # In discover mode, log compactly but don't write to JSONL
            if self.discover:
                direction = 'OUT' if rec.get('isOutbound') else 'IN '
                room = (rec.get('roomName') or '?')[:30]
                sender = (rec.get('senderName') or '?')[:25]
                firm = (rec.get('firmName') or '?')[:20]
                log(f'  [DISCOVER] [{direction}] {room:<30} {sender:<25} ({firm})')
                return

            # Apply filter
            if not self.filter.should_keep(rec):
                return

            self.seen_ids.add(msg_id)
            if append_jsonl(rec):
                self._new_msgs_since_rebuild += 1
                # Compact log line
                direction = 'OUT' if rec.get('isOutbound') else 'IN '
                room = (rec.get('roomName') or '?')[:25]
                sender = (rec.get('senderName') or '?')[:20]
                body_preview = (rec.get('body') or '')[:60].replace('\n', ' ')
                log(f'  [{direction}] {room:<25} {sender:<20} : {body_preview}')

    def _maybe_rebuild_xlsx(self):
        """Rebuild Excel and save discovery if enough time has passed."""
        if time.time() - self._last_xlsx_rebuild < XLSX_REBUILD_INTERVAL:
            return
        if self._new_msgs_since_rebuild > 0:
            rebuild_xlsx()
            self._new_msgs_since_rebuild = 0
        # Save discovery snapshot periodically (cheap, JSON dump)
        if self._discovery:
            self._save_discovery()
        self._last_xlsx_rebuild = time.time()

    def run(self):
        """Main event loop. Blocks until stop() called or fatal error."""
        if not self._connect():
            return False
        if not self._open_msgscrape():
            return False
        if not self._subscribe():
            return False

        import blpapi
        self._running = True
        log('Listening for IB messages. Ctrl-C to stop.')

        try:
            while self._running:
                ev = self._session.nextEvent(1000)  # 1s timeout
                ev_type = ev.eventType()

                if ev_type in (blpapi.Event.SUBSCRIPTION_DATA,
                               blpapi.Event.PARTIAL_RESPONSE,
                               blpapi.Event.RESPONSE):
                    for msg in ev:
                        try:
                            self._process_message(msg)
                        except Exception as e:
                            log(f'  ERROR processing message: {e}')

                elif ev_type == blpapi.Event.SUBSCRIPTION_STATUS:
                    for msg in ev:
                        log(f'  subscription status: {msg}')

                elif ev_type == blpapi.Event.SESSION_STATUS:
                    for msg in ev:
                        msg_type = msg.messageType()
                        log(f'  session status: {msg_type}')
                        # If session was terminated, exit so the outer loop can reconnect
                        if str(msg_type) in ('SessionTerminated', 'SessionStartupFailure'):
                            log('  session terminated — will attempt reconnect')
                            self._running = False

                # Periodic Excel rebuild
                self._maybe_rebuild_xlsx()

        except KeyboardInterrupt:
            log('Interrupted by user')
            self._running = False
        finally:
            try:
                self._session.stop()
            except Exception:
                pass
            # Final Excel rebuild and discovery save on shutdown
            if self._new_msgs_since_rebuild > 0:
                rebuild_xlsx()
            if self._discovery:
                self._save_discovery()
                log(f'Discovery snapshot saved to {SEEN_PATH.name} ({len(self._discovery)} unique sender+room combos)')
        return True

    def stop(self):
        self._running = False


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Continuous IB chat scraper')
    parser.add_argument('--once', action='store_true',
                        help='Run for 60 seconds then exit (testing)')
    parser.add_argument('--rebuild-xlsx', action='store_true',
                        help='Just rebuild Excel from JSONL and exit')
    parser.add_argument('--config', type=str, default=None,
                        help='Path to filter config JSON (default: scrape_filters.json next to script)')
    parser.add_argument('--no-filter', action='store_true',
                        help='Capture ALL messages regardless of config (legacy behavior)')
    parser.add_argument('--discover', action='store_true',
                        help='Log every sender/room seen but DO NOT save messages. Use to find UUIDs/names/firms/rooms for your filter config.')
    parser.add_argument('--reconnect-delay', type=int, default=30,
                        help='Seconds to wait between reconnect attempts on disconnect')
    args = parser.parse_args()

    if args.rebuild_xlsx:
        rebuild_xlsx()
        return

    # Load filter config
    config_path = Path(args.config) if args.config else FILTER_PATH
    config = load_filter_config(config_path)
    if args.no_filter:
        config['capture_all'] = True
        log('--no-filter: ignoring filter config, capturing everything')

    chat_filter = ChatFilter(config)
    log(f'IB Chat Scraper starting')
    log(f'  Filter: {chat_filter.describe()}')
    log(f'  Config: {config_path}')
    log(f'  JSONL:  {JSONL_PATH}')
    log(f'  XLSX:   {XLSX_PATH}')
    log(f'  Log:    {LOG_PATH}')

    if args.discover:
        log('  DISCOVER MODE: messages will NOT be saved to JSONL.')
        log(f'  Sender/room metadata will be saved to {SEEN_PATH.name} so you can pick filters.')

    # Wrapper loop with auto-reconnect
    while True:
        scraper = IBScraper(chat_filter=chat_filter, discover_mode=args.discover)
        try:
            if args.once:
                # One-shot mode: run 60s then exit
                t = threading.Thread(target=scraper.run, daemon=True)
                t.start()
                time.sleep(60)
                scraper.stop()
                t.join(timeout=5)
                break
            else:
                ok = scraper.run()
                if ok is False:
                    log(f'Initial connect/setup failed — retrying in {args.reconnect_delay}s')
                    time.sleep(args.reconnect_delay)
                else:
                    log(f'Session ended — reconnecting in {args.reconnect_delay}s')
                    time.sleep(args.reconnect_delay)
        except KeyboardInterrupt:
            log('Shutting down')
            break
        except Exception as e:
            log(f'Unexpected error: {e}')
            time.sleep(args.reconnect_delay)


if __name__ == '__main__':
    main()
