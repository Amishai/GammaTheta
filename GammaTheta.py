#!/usr/bin/env python3
"""
FX Gamma Trading Dashboard

Core formula: |Theta + Delta * T/N_roll| / Gamma
Output: Normalized to 1-5 scale (lower = cheaper gamma)
"""

import numpy as np
import pandas as pd
from scipy.stats import norm
from scipy.optimize import brentq
import json
import os

# === BLACK-SCHOLES ===

def d1(S, K, T, sigma, r_d, r_f):
    if T <= 0 or sigma <= 0: return 0.0
    return (np.log(S / K) + (r_d - r_f + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))

def d2(S, K, T, sigma, r_d, r_f):
    return d1(S, K, T, sigma, r_d, r_f) - sigma * np.sqrt(T)

def bs_delta(S, K, T, sigma, r_d, r_f, is_call):
    if T <= 0: return 0.0
    d1_val = d1(S, K, T, sigma, r_d, r_f)
    df_f = np.exp(-r_f * T)
    return df_f * norm.cdf(d1_val) if is_call else df_f * (norm.cdf(d1_val) - 1)

def bs_gamma(S, K, T, sigma, r_d, r_f):
    if T <= 0 or sigma <= 0: return 0.0
    d1_val = d1(S, K, T, sigma, r_d, r_f)
    df_f = np.exp(-r_f * T)
    return df_f * norm.pdf(d1_val) / (S * sigma * np.sqrt(T))

def bs_theta(S, K, T, sigma, r_d, r_f, is_call):
    if T <= 0: return 0.0
    d1_val, d2_val = d1(S, K, T, sigma, r_d, r_f), d2(S, K, T, sigma, r_d, r_f)
    df_d, df_f = np.exp(-r_d * T), np.exp(-r_f * T)
    term1 = -S * df_f * norm.pdf(d1_val) * sigma / (2 * np.sqrt(T))
    if is_call:
        term2 = r_f * S * df_f * norm.cdf(d1_val) - r_d * K * df_d * norm.cdf(d2_val)
    else:
        term2 = -r_f * S * df_f * norm.cdf(-d1_val) + r_d * K * df_d * norm.cdf(-d2_val)
    return (term1 + term2) / 365

def bs_vega(S, K, T, sigma, r_d, r_f):
    if T <= 0: return 0.0
    d1_val = d1(S, K, T, sigma, r_d, r_f)
    df_f = np.exp(-r_f * T)
    return S * df_f * norm.pdf(d1_val) * np.sqrt(T) * 0.01

# === STRIKE FROM DELTA ===

def strike_from_delta_call(F, T, sigma, target_delta, df_f):
    def obj(K):
        d1_val = (np.log(F / K) + 0.5 * sigma**2 * T) / (sigma * np.sqrt(T))
        return df_f * norm.cdf(d1_val) - target_delta
    try: return brentq(obj, F * 0.5, F * 2.0)
    except: return F

def strike_from_delta_put(F, T, sigma, target_delta, df_f):
    def obj(K):
        d1_val = (np.log(F / K) + 0.5 * sigma**2 * T) / (sigma * np.sqrt(T))
        return df_f * (norm.cdf(d1_val) - 1) - target_delta
    try: return brentq(obj, F * 0.5, F * 2.0)
    except: return F

def atm_dns_strike(F, T, sigma):
    return F * np.exp(0.5 * sigma**2 * T)

# === SMILE ===

def solve_smile(atm_vol, rr25, fly25, rr10, fly10):
    atm = atm_vol / 100
    return {
        'ATM': atm,
        '25C': atm + fly25/100 + 0.5 * rr25/100,
        '25P': atm + fly25/100 - 0.5 * rr25/100,
        '10C': atm + fly10/100 + 0.5 * rr10/100,
        '10P': atm + fly10/100 - 0.5 * rr10/100
    }

def interp_vol(smile, delta, is_call):
    if delta == 0: return smile['ATM']
    ad = abs(delta)
    if is_call:
        if ad <= 0.10: return smile['10C']
        elif ad <= 0.25: return smile['10C'] + (ad - 0.10) / 0.15 * (smile['25C'] - smile['10C'])
        else: return smile['25C'] + (ad - 0.25) / 0.25 * (smile['ATM'] - smile['25C'])
    else:
        if ad <= 0.10: return smile['10P']
        elif ad <= 0.25: return smile['10P'] + (ad - 0.10) / 0.15 * (smile['25P'] - smile['10P'])
        else: return smile['25P'] + (ad - 0.25) / 0.25 * (smile['ATM'] - smile['25P'])

# === EXCEL I/O ===

def create_template(fp):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta

    wb = Workbook()
    hf, hfill = Font(bold=True, color='FFFFFF'), PatternFill('solid', fgColor='1F4E79')
    inf, infill = Font(color='0000FF'), PatternFill('solid', fgColor='D6EAF8')
    bord = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws = wb.active
    ws.title = 'Parameters'
    for i, h in enumerate(['Parameter', 'Value', 'Description'], 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.border = hf, hfill, bord

    for i, (n, v, d) in enumerate([
        ('Spot', 1.0850, 'Spot rate'),
        ('TermsRate', 0.045, 'Terms/quote currency rate'),
        ('BaseRate', 0.025, 'Base currency rate')
    ], 2):
        ws.cell(row=i, column=1, value=n).border = bord
        c = ws.cell(row=i, column=2, value=v)
        c.font, c.fill, c.border = inf, infill, bord
        ws.cell(row=i, column=3, value=d).border = bord
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 15, 12, 35

    ws2 = wb.create_sheet('VolSurface')
    for i, h in enumerate(['Tenor', 'T_Years', 'ATM', 'RR25', 'RR10', 'FLY25', 'FLY10', 'FwdPts'], 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font, c.fill, c.border = hf, hfill, bord

    data = [
        ('O/N', 1/365, 7.50, -0.30, -0.60, 0.15, 0.40, 0.59),
        ('1W', 7/365, 7.80, -0.35, -0.70, 0.18, 0.45, 4.16),
        ('2W', 14/365, 8.00, -0.40, -0.80, 0.20, 0.50, 8.33),
        ('1M', 1/12, 8.20, -0.50, -1.00, 0.25, 0.60, 18.10),
        ('2M', 2/12, 8.50, -0.60, -1.20, 0.30, 0.70, 36.23),
        ('3M', 3/12, 8.80, -0.70, -1.40, 0.35, 0.80, 54.39),
        ('6M', 6/12, 9.20, -0.80, -1.60, 0.40, 0.95, 109.04),
        ('9M', 9/12, 9.50, -0.85, -1.70, 0.42, 1.05, 163.98),
        ('1Y', 1.0, 9.80, -0.90, -1.80, 0.45, 1.15, 219.18),
        ('2Y', 2.0, 10.20, -1.00, -2.00, 0.50, 1.30, 442.80),
    ]
    for ri, row in enumerate(data, 2):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.border = bord
            if ci >= 3: c.font, c.fill = inf, infill
    for i, w in enumerate([8, 10, 10, 10, 10, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet('Positions')
    for i, h in enumerate(['Strike', 'Expiry', 'Vol', 'Notional', 'Type'], 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font, c.fill, c.border = hf, hfill, bord

    ws3.cell(row=2, column=1, value='# Strike: option strike price')
    ws3.cell(row=3, column=1, value='# Expiry: YYYY-MM-DD format')
    ws3.cell(row=4, column=1, value='# Vol: implied vol in %')
    ws3.cell(row=5, column=1, value='# Notional: $M (negative for short)')
    ws3.cell(row=6, column=1, value='# Type: C or P')
    for i in range(2, 7):
        ws3.cell(row=i, column=1).font = Font(italic=True, color='888888')

    today = datetime.now()
    sample_positions = [
        (1.0900, (today + timedelta(days=90)).strftime('%Y-%m-%d'), 8.5, 10, 'C'),
        (1.0750, (today + timedelta(days=45)).strftime('%Y-%m-%d'), 9.2, -5, 'P'),
        (1.1000, (today + timedelta(days=180)).strftime('%Y-%m-%d'), 7.8, 15, 'C'),
        (1.0800, (today + timedelta(days=60)).strftime('%Y-%m-%d'), 8.8, -8, 'C'),
        (1.0650, (today + timedelta(days=14)).strftime('%Y-%m-%d'), 10.1, 5, 'P'),
        (1.1100, (today + timedelta(days=270)).strftime('%Y-%m-%d'), 7.2, 20, 'C'),
    ]
    for ri, row in enumerate(sample_positions, 8):
        for ci, v in enumerate(row, 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.border = bord
            c.font, c.fill = inf, infill

    for i, w in enumerate([12, 12, 8, 12, 8], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(fp)

def load_data(fp):
    from datetime import datetime
    p = pd.read_excel(fp, sheet_name='Parameters')
    params = dict(zip(p['Parameter'], p['Value']))
    v = pd.read_excel(fp, sheet_name='VolSurface')

    S = params['Spot']
    r_d = params['TermsRate']

    # Handle both old (7-col) and new (8-col with FwdPts) formats
    if v.shape[1] >= 8:
        v.columns = ['tenor', 'T_years', 'atm', 'rr25', 'rr10', 'fly25', 'fly10', 'fwd_pts']
    else:
        v.columns = ['tenor', 'T_years', 'atm', 'rr25', 'rr10', 'fly25', 'fly10']
        # Fall back to rate-derived forward points
        r_f_flat = params['BaseRate']
        v['fwd_pts'] = v['T_years'].apply(lambda T: (S * np.exp((r_d - r_f_flat) * T) - S) * 10000)

    # Derive forward and implied r_f for each tenor from actual fwd pts
    v['forward'] = S + v['fwd_pts'] / 10000
    v['r_f_implied'] = v.apply(lambda r: r_d - np.log(r['forward'] / S) / r['T_years'] if r['T_years'] > 0 else params.get('BaseRate', 0.025), axis=1)

    # Build forward curve for interpolation (used by portfolio positions)
    fwd_curve = v[['T_years', 'forward', 'r_f_implied']].to_dict('records')

    positions = []
    try:
        pos_df = pd.read_excel(fp, sheet_name='Positions')
        pos_df.columns = ['strike', 'expiry', 'vol', 'notional', 'type']
        pos_df = pos_df[pos_df['strike'].apply(lambda x: isinstance(x, (int, float)) and not pd.isna(x))]
        today = datetime.now()
        for _, row in pos_df.iterrows():
            try:
                expiry = pd.to_datetime(row['expiry'])
                T = max(1/365, (expiry - pd.Timestamp(today)).days / 365)
                positions.append({
                    'strike': float(row['strike']),
                    'expiry': expiry.strftime('%Y-%m-%d'),
                    'vol': float(row['vol']),
                    'notional': float(row['notional']),
                    'type': str(row['type']).upper().strip(),
                    'T': T,
                    'days': max(1, int(T * 365))
                })
            except:
                continue
    except:
        pass

    # T/N roll cost = negative of O/N fwd pts
    # (positive fwd pts = base at premium = costs to roll short base hedge)
    if 'TN_Roll_Pips' in params:
        tn_roll = params['TN_Roll_Pips']
    else:
        on_row = v[v['T_years'] == v['T_years'].min()].iloc[0]
        tn_roll = -on_row['fwd_pts']

    return {'spot': S, 'r_terms': r_d, 'r_base': params.get('BaseRate', 0.025),
            'tn_roll': tn_roll, 'vol_surface': v, 'positions': positions,
            'fwd_curve': fwd_curve}

# === BUILD SURFACE ===

def build_surface(mkt):
    S, r_d, tn = mkt['spot'], mkt['r_terms'], mkt['tn_roll']
    notional = 1_000_000

    deltas = [('10P', -0.10, False), ('15P', -0.15, False), ('20P', -0.20, False), ('25P', -0.25, False),
              ('30P', -0.30, False), ('35P', -0.35, False), ('40P', -0.40, False), ('45P', -0.45, False),
              ('ATM', 0.0, None),
              ('45C', 0.45, True), ('40C', 0.40, True), ('35C', 0.35, True), ('30C', 0.30, True),
              ('25C', 0.25, True), ('20C', 0.20, True), ('15C', 0.15, True), ('10C', 0.10, True)]

    rows = []
    fwd_points_list = []

    for _, r in mkt['vol_surface'].iterrows():
        T = r['T_years']
        F = r['forward']
        r_f = r['r_f_implied']
        df_f = np.exp(-r_f * T)
        smile = solve_smile(r['atm'], r['rr25'], r['fly25'], r['rr10'], r['fly10'])

        fwd_pts = r['fwd_pts']
        days = max(1, round(T * 365))
        fwd_points_list.append({'tenor': r['tenor'], 'T_years': T, 'days': days, 'forward': F, 'fwd_points': fwd_pts})

        for dname, dval, is_call in deltas:
            if dval == 0:
                sigma = smile['ATM']
                K = atm_dns_strike(F, T, sigma)
                theta = (bs_theta(S, K, T, sigma, r_d, r_f, True) + bs_theta(S, K, T, sigma, r_d, r_f, False)) / 2
                delta = round(bs_delta(S, K, T, sigma, r_d, r_f, True), 4)
                is_call_flag = True
            else:
                sigma = interp_vol(smile, dval, is_call)
                K = strike_from_delta_call(F, T, sigma, abs(dval), df_f) if is_call else strike_from_delta_put(F, T, sigma, dval, df_f)
                theta = bs_theta(S, K, T, sigma, r_d, r_f, is_call)
                delta = bs_delta(S, K, T, sigma, r_d, r_f, is_call)
                is_call_flag = is_call

            gamma = bs_gamma(S, K, T, sigma, r_d, r_f)
            vega = bs_vega(S, K, T, sigma, r_d, r_f)

            # Core formula: |Theta + Delta * T/N_roll| / Gamma
            roll = delta * (tn / 10000)
            cost = abs(theta + roll) / gamma if gamma > 1e-12 else 0

            theta_daily = theta * notional
            total_decay = theta_daily * days

            rows.append({
                'tenor': r['tenor'], 'T_years': T, 'days': days, 'delta_label': dname, 'delta_val': dval,
                'strike': K, 'vol': sigma * 100, 'delta': delta,
                'gamma': gamma * S * 0.01 * notional / 1_000_000,
                'theta': theta * notional / 1000,
                'total_decay': total_decay / 1000,
                'vega': vega * notional / 1000,
                'gamma_cost_raw': cost
            })

    df = pd.DataFrame(rows)
    raw = df['gamma_cost_raw'].values
    p5 = float(np.percentile(raw[raw > 0], 5))
    p95 = float(np.percentile(raw[raw > 0], 95))
    df['richness'] = df['gamma_cost_raw'].apply(lambda x: max(1, min(5, 1 + (x - p5) / (p95 - p5) * 4)) if x > 0 else 1)

    fwd_df = pd.DataFrame(fwd_points_list)
    return df, fwd_df, p5, p95

# === HTML DASHBOARD ===

def create_dashboard(df, fwd_df, mkt, p5, p95, output='fx_gamma_trading.html'):
    from datetime import datetime

    tenors = df['tenor'].unique().tolist()
    deltas = ['10P', '15P', '20P', '25P', '30P', '35P', '40P', '45P', 'ATM', '45C', '40C', '35C', '30C', '25C', '20C', '15C', '10C']
    dlabels = ['10\u0394 P', '15\u0394 P', '20\u0394 P', '25\u0394 P', '30\u0394 P', '35\u0394 P', '40\u0394 P', '45\u0394 P', 'ATM', '45\u0394 C', '40\u0394 C', '35\u0394 C', '30\u0394 C', '25\u0394 C', '20\u0394 C', '15\u0394 C', '10\u0394 C']

    matrix_richness = [[float(df[(df['tenor']==t) & (df['delta_label']==d)]['richness'].values[0])
               for d in deltas] for t in tenors]
    matrix_theta = [[float(df[(df['tenor']==t) & (df['delta_label']==d)]['theta'].values[0])
               for d in deltas] for t in tenors]
    matrix_gamma = [[float(df[(df['tenor']==t) & (df['delta_label']==d)]['gamma'].values[0])
               for d in deltas] for t in tenors]
    matrix_vega = [[float(df[(df['tenor']==t) & (df['delta_label']==d)]['vega'].values[0])
               for d in deltas] for t in tenors]
    matrix_strikes = [[round(float(df[(df['tenor']==t) & (df['delta_label']==d)]['strike'].values[0]), 3)
               for d in deltas] for t in tenors]
    matrix_delta = [[float(df[(df['tenor']==t) & (df['delta_label']==d)]['delta'].values[0])
               for d in deltas] for t in tenors]

    sdata = df.to_dict('records')
    fwd_data = fwd_df.to_dict('records')

    tenor_order = fwd_df.sort_values('T_years')['tenor'].tolist()

    cheapest_gamma = []
    richest_gamma = []
    for t in tenor_order:
        t_df = df[df['tenor'] == t]
        cheapest = t_df.loc[t_df['richness'].idxmin()]
        richest = t_df.loc[t_df['richness'].idxmax()]
        cheapest_gamma.append({
            'tenor': t, 'delta': cheapest['delta_label'], 'score': cheapest['richness'],
            'theta': cheapest['theta'], 'gamma': cheapest['gamma'], 'vega': cheapest['vega'],
            'total_decay': cheapest['total_decay']
        })
        richest_gamma.append({
            'tenor': t, 'delta': richest['delta_label'], 'score': richest['richness'],
            'theta': richest['theta'], 'gamma': richest['gamma'], 'vega': richest['vega'],
            'total_decay': richest['total_decay']
        })

    fwd_table = '<table><tr><th>Tenor</th><th>Days</th><th>Forward</th><th>Fwd Pts</th></tr>'
    for f in fwd_data:
        fwd_table += f'<tr><td>{f["tenor"]}</td><td>{f["days"]}</td><td>{f["forward"]:.5f}</td><td>{f["fwd_points"]:.2f}</td></tr>'
    fwd_table += '</table>'

    tenor_days = {r['tenor']: r['days'] for r in fwd_data}
    positions = mkt.get('positions', [])
    today_str = datetime.now().strftime('%Y-%m-%d')

    on_fwd_pts = fwd_df.loc[fwd_df['T_years'].idxmin(), 'fwd_points'] if len(fwd_df) > 0 else 0

    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>FX Gamma Trading</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
*{{box-sizing:border-box}}body{{font-family:'Segoe UI',Arial,sans-serif;margin:0;padding:15px;background:#2d2d2d;color:#e0e0e0}}
h1{{text-align:center;color:#90caf9;margin:10px 0 15px}}h2{{color:#90caf9;margin:0 0 15px;font-size:18px}}h3{{color:#90caf9;margin:15px 0 10px;font-size:14px}}
.grid{{display:grid;grid-template-columns:1fr 420px;gap:15px;max-width:1600px;margin:0 auto}}
.card{{background:#3d3d3d;padding:15px;border-radius:8px;box-shadow:0 2px 4px rgba(0,0,0,0.3)}}
.market-bar{{display:flex;gap:20px;justify-content:center;margin-bottom:15px;padding:10px;background:#3d3d3d;border-radius:8px;flex-wrap:wrap}}
.market-item{{text-align:center}}.market-item .label{{font-size:11px;color:#aaa}}.market-item .value{{font-size:16px;font-weight:bold;color:#90caf9}}
.legend{{display:flex;justify-content:center;gap:20px;margin:10px 0;font-size:12px}}
.legend-item{{display:flex;align-items:center;gap:5px}}.legend-color{{width:16px;height:16px;border-radius:3px}}
table{{width:100%;border-collapse:collapse;font-size:11px;margin-bottom:10px}}th,td{{padding:5px 6px;text-align:right;border:1px solid #555}}
th{{background:#1F4E79;color:white}}td:first-child{{text-align:left;font-weight:600}}.cheap{{background:#1a3a4a}}.rich{{background:#4a2a2a}}
.btn{{padding:8px 16px;border:none;border-radius:4px;cursor:pointer;font-weight:600;margin-right:8px}}
.btn-primary{{background:#1F4E79;color:white}}.btn-primary:hover{{background:#2a5f8f}}
.btn-toggle{{padding:6px 12px;border:1px solid #555;border-radius:4px;cursor:pointer;font-weight:600;margin-right:4px;background:#2d2d2d;color:#e0e0e0}}
.btn-toggle.active{{background:#1F4E79;color:white;border-color:#1F4E79}}
.greeks-result{{background:#2d2d2d;padding:12px;border-radius:6px;margin-top:10px}}
.greeks-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;text-align:center}}
.greek-box{{padding:8px;background:#3d3d3d;border-radius:4px}}.greek-label{{font-size:10px;color:#aaa}}
.greek-value{{font-size:14px;font-weight:bold;color:#90caf9}}
.help{{font-size:11px;color:#aaa;margin-top:8px}}
.unit{{font-size:9px;color:#888}}
select{{padding:6px;background:#3d3d3d;color:#e0e0e0;border:1px solid #555;border-radius:4px}}
.notional-input{{width:80px;padding:6px;background:#2d2d2d;color:#90caf9;border:1px solid #555;border-radius:4px;font-size:14px;font-weight:bold;text-align:center}}
.tab-bar{{display:flex;gap:0;margin-bottom:15px;max-width:1600px;margin-left:auto;margin-right:auto}}
.tab-btn{{padding:12px 24px;border:none;background:#3d3d3d;color:#aaa;font-size:14px;font-weight:600;cursor:pointer;border-bottom:3px solid transparent;transition:all 0.2s}}
.tab-btn:hover{{color:#e0e0e0;background:#454545}}
.tab-btn.active{{color:#90caf9;border-bottom-color:#90caf9;background:#3d3d3d}}
.tab-content{{display:none}}.tab-content.active{{display:block}}
.portfolio-grid{{display:grid;grid-template-columns:1fr 1fr;gap:15px;max-width:1600px;margin:0 auto}}
.summary-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:15px}}
.summary-box{{background:#2d2d2d;padding:12px;border-radius:6px;text-align:center}}
.summary-label{{font-size:11px;color:#aaa;margin-bottom:4px}}
.summary-value{{font-size:18px;font-weight:bold;color:#90caf9}}
.modal{{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.7);z-index:1000}}
.modal.show{{display:flex;align-items:center;justify-content:center}}
.modal-content{{background:#3d3d3d;padding:20px;border-radius:8px;max-width:800px;max-height:80vh;overflow-y:auto}}
.modal-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:15px}}
.modal-close{{background:none;border:none;color:#e0e0e0;font-size:24px;cursor:pointer}}
.no-positions{{text-align:center;padding:40px;color:#aaa}}
.richness-badge{{display:inline-block;padding:2px 6px;border-radius:3px;font-weight:bold;font-size:11px;color:#000}}
</style></head><body>
<h1>FX Gamma Trading Dashboard</h1>
<div class="market-bar">
<div class="market-item"><div class="label">Spot</div><div class="value">{mkt['spot']:.4f}</div></div>
<div class="market-item"><div class="label">Terms Rate</div><div class="value">{mkt['r_terms']*100:.2f}%</div></div>
<div class="market-item"><div class="label">Base Rate</div><div class="value">{mkt['r_base']*100:.2f}%</div></div>
<div class="market-item"><div class="label">O/N Fwd Pts</div><div class="value">{on_fwd_pts:.2f} pips</div></div>
<div class="market-item"><div class="label">Notional ($M)</div><input type="number" id="notional" class="notional-input" value="1" min="0.1" step="0.1" onchange="updateNotional()"></div>
</div>

<div class="tab-bar">
<button class="tab-btn active" onclick="showTab('surface')">Surface Analysis</button>
<button class="tab-btn" onclick="showTab('portfolio')">Portfolio Analysis</button>
</div>

<!-- TAB 1: Surface Analysis -->
<div id="tab-surface" class="tab-content active">
<div class="grid"><div>
<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:10px">
<h2 style="margin:0" id="heatmap-title">Richness Score</h2>
<div>
<button class="btn-toggle active" onclick="setHeatmapView('richness')" id="hm-btn-richness">Richness</button>
<button class="btn-toggle" onclick="setHeatmapView('theta')" id="hm-btn-theta">Theta</button>
<button class="btn-toggle" onclick="setHeatmapView('vega')" id="hm-btn-vega">Vega</button>
</div>
</div>
<div id="legend-richness" class="legend">
<div class="legend-item"><div class="legend-color" style="background:#4caf50"></div><span>1 (Cheap)</span></div>
<div class="legend-item"><div class="legend-color" style="background:#fff59d"></div><span>3 (Fair)</span></div>
<div class="legend-item"><div class="legend-color" style="background:#f44336"></div><span>5 (Rich)</span></div>
</div>
<div id="legend-theta" class="legend" style="display:none">
<div class="legend-item"><div class="legend-color" style="background:#ef5350"></div><span>High Decay</span></div>
<div class="legend-item"><div class="legend-color" style="background:#1a237e"></div><span>Low Decay</span></div>
</div>
<div id="legend-vega" class="legend" style="display:none">
<div class="legend-item"><div class="legend-color" style="background:#0d47a1"></div><span>Low Vega</span></div>
<div class="legend-item"><div class="legend-color" style="background:#b71c1c"></div><span>High Vega</span></div>
</div>
<div id="heatmap"></div>
<div class="help">Richness = |Theta + Delta*T/N Roll| / Gamma, normalized 1-5 | Lower = cheaper gamma</div></div>

<div class="card" style="margin-top:15px"><h2>Richness by Tenor</h2>
<h3>Cheapest Gamma per Tenor (Buy)</h3>
<div id="cheap-table"></div>
<h3>Richest Gamma per Tenor (Sell)</h3>
<div id="rich-table"></div>
<div class="help">Total Decay = Daily Theta x Days to Expiry</div>
</div></div>

<div>
<div class="card"><h2>Forward Points</h2>
{fwd_table}
<div class="help">Fwd Pts in pips (terms currency)</div></div>

<div class="card" style="margin-top:15px"><h2>Greeks Calculator</h2>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
<div><label style="font-size:12px">Tenor</label><select id="calc-t" style="width:100%"></select></div>
<div><label style="font-size:12px">Delta</label><select id="calc-d" style="width:100%"></select></div>
</div><button class="btn btn-primary" onclick="calcGreeks()" style="margin-top:10px">Calculate</button>
<div id="calc-result"></div></div>
</div></div>
</div>

<!-- TAB 2: Portfolio Analysis -->
<div id="tab-portfolio" class="tab-content">
<div id="portfolio-content">
<div class="summary-grid" style="max-width:1600px;margin:0 auto 15px">
<div class="summary-box"><div class="summary-label">Total Gamma ($M)</div><div class="summary-value" id="sum-gamma">-</div></div>
<div class="summary-box"><div class="summary-label">Total Theta ($K/day)</div><div class="summary-value" id="sum-theta">-</div></div>
<div class="summary-box"><div class="summary-label">Wtd Avg Richness</div><div class="summary-value" id="sum-richness">-</div></div>
<div class="summary-box"><div class="summary-label">Total Vega ($K)</div><div class="summary-value" id="sum-vega">-</div></div>
<div class="summary-box"><div class="summary-label">Projected Total Decay ($K)</div><div class="summary-value" id="sum-decay">-</div></div>
</div>

<div class="portfolio-grid">
<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:10px">
<h2 style="margin:0" id="port-heatmap-title">Portfolio Richness (1-5)</h2>
<div>
<button class="btn-toggle active" onclick="setPortHeatmapView('richness')" id="port-btn-richness">Richness</button>
<button class="btn-toggle" onclick="setPortHeatmapView('vega')" id="port-btn-vega">Vega</button>
</div>
</div>
<div id="port-heatmap"></div>
<div class="help">Click on a cell to see positions in that bucket. Gamma-weighted avg richness (1-5 scale).</div>
</div>

<div class="card">
<h2>Position Details ({len(positions)} positions from Excel)</h2>
<div id="positions-table" style="max-height:400px;overflow-y:auto"></div>
<div class="help" style="margin-top:10px">Edit positions in the "Positions" sheet of fx_gamma_inputs.xlsx and re-run the script.</div>
</div>

<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:10px">
<h2 style="margin:0">Inefficient Positions</h2>
<div>
<button class="btn-toggle active" onclick="setIneffView('long')" id="ineff-btn-long">Long (Rich Gamma)</button>
<button class="btn-toggle" onclick="setIneffView('short')" id="ineff-btn-short">Short (Cheap Gamma)</button>
</div>
</div>
<div id="ineff-table" style="max-height:300px;overflow-y:auto"></div>
<div class="help" style="margin-top:10px">Longs with highest richness (overpaying for gamma) vs Shorts with lowest richness (selling cheap gamma).</div>
</div>
</div>

<div style="max-width:1600px;margin:15px auto 0">
<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:10px">
<h2 style="margin:0">Portfolio Greeks Over Time</h2>
<div>
<button class="btn-toggle active" onclick="setTimeView('gamma')" id="time-btn-gamma">Gamma</button>
<button class="btn-toggle" onclick="setTimeView('theta')" id="time-btn-theta">Theta</button>
<button class="btn-toggle" onclick="setTimeView('vega')" id="time-btn-vega">Vega</button>
<button class="btn-toggle" onclick="setTimeView('cumDecay')" id="time-btn-cumDecay">Cumulative Decay</button>
</div>
</div>
<div id="time-chart" style="height:350px"></div>
<div class="help">Projects portfolio Greeks forward day-by-day as options age and expire. Cum. Decay = running total of daily theta paid/received.</div>
</div>
</div>

</div>
</div>
<div id="drill-modal" class="modal" onclick="if(event.target===this)closeDrillDown()">
<div class="modal-content">
<div class="modal-header">
<h2 id="drill-title" style="margin:0">Positions</h2>
<button class="modal-close" onclick="closeDrillDown()">&times;</button>
</div>
<div id="drill-content"></div>
</div>
</div>

<script>
var S={json.dumps(sdata)},T={json.dumps(tenors)},D={json.dumps(deltas)},DL={json.dumps(dlabels, ensure_ascii=False)};
var M_richness={json.dumps(matrix_richness)};
var M_theta={json.dumps(matrix_theta)};
var M_gamma={json.dumps(matrix_gamma)};
var M_vega={json.dumps(matrix_vega)};
var M_strikes={json.dumps(matrix_strikes)};
var M_delta={json.dumps(matrix_delta)};
var tenorDays={json.dumps(tenor_days)};
var mkt={{spot:{mkt['spot']},r_d:{mkt['r_terms']},r_f:{mkt['r_base']},tn_roll:{mkt['tn_roll']}}};
var fwdCurve={json.dumps(mkt['fwd_curve'])};
var excelPositions={json.dumps(positions)};
var todayStr="{today_str}";

// Surface normalization bounds - positions use these same bounds for consistent 1-5 scale
var surfaceP5={p5};
var surfaceP95={p95};

var cs_richness=[[0,'#4caf50'],[0.5,'#fff59d'],[1,'#f44336']];
var cs_vega=[[0,'#0d47a1'],[0.5,'#fff59d'],[1,'#b71c1c']];

var greekData={{gamma:{{cheap:{json.dumps(cheapest_gamma)},rich:{json.dumps(richest_gamma)}}}}};

var currentHeatmap='richness';
var currentPortHeatmap='richness';
var portfolioData=null;

// Normalize raw gamma cost to 1-5 using the surface's percentile bounds
function rawToRichness(raw){{
    if(raw<=0) return 1;
    return Math.max(1,Math.min(5, 1+(raw-surfaceP5)/(surfaceP95-surfaceP5)*4));
}}

// Richness color: 1=green, 3=yellow, 5=red
function richnessColor(score){{
    if(score<=3){{
        var t=(score-1)/2;
        var r=Math.round(76+(255-76)*t), g=Math.round(175+(245-175)*t), b=Math.round(80+(157-80)*t);
        return 'rgb('+r+','+g+','+b+')';
    }}else{{
        var t=(score-3)/2;
        var r=Math.round(255-(255-244)*t), g=Math.round(245-(245-67)*t), b=Math.round(157-(157-54)*t);
        return 'rgb('+r+','+g+','+b+')';
    }}
}}

function showTab(tab){{
    document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
    document.querySelectorAll('.tab-content').forEach(c=>c.classList.remove('active'));
    document.querySelector('.tab-btn[onclick*="'+tab+'"]').classList.add('active');
    document.getElementById('tab-'+tab).classList.add('active');
    if(tab==='surface') renderHeatmap();
    if(tab==='portfolio'){{if(!portfolioData) initPortfolio(); renderPortHeatmap();}}
}}

function getNotional(){{return parseFloat(document.getElementById('notional').value)||1;}}

function updateNotional(){{
    renderTables(); renderHeatmap();
    if(document.getElementById('calc-result').innerHTML) calcGreeks();
}}

function setHeatmapView(view){{
    currentHeatmap=view;
    ['richness','theta','vega'].forEach(v=>{{
        document.getElementById('hm-btn-'+v).className='btn-toggle'+(v===view?' active':'');
        document.getElementById('legend-'+v).style.display=(v===view?'flex':'none');
    }});
    var titles={{richness:'Richness Score',theta:'Theta Decay ($K/day)',vega:'Vega Surface ($K)'}};
    document.getElementById('heatmap-title').textContent=titles[view];
    renderHeatmap();
}}

function renderHeatmap(){{
    var n=getNotional();
    var M,cs,zmin,zmax,barTitle;
    if(currentHeatmap==='richness'){{
        M=M_richness; cs=cs_richness; zmin=1; zmax=5; barTitle='Score';
    }}else if(currentHeatmap==='theta'){{
        M=M_theta.map(r=>r.map(v=>Math.abs(v*n)));
        cs=[[0,'#1a237e'],[0.5,'#fff59d'],[1,'#ef5350']];
        var flat=M.flat(); zmin=Math.min(...flat); zmax=Math.max(...flat); barTitle='$K/day';
    }}else{{
        M=M_vega.map(r=>r.map(v=>Math.abs(v*n))); cs=cs_vega;
        var flat=M.flat(); zmin=Math.min(...flat); zmax=Math.max(...flat); barTitle='$K';
    }}
    var txt=M.map(r=>r.map(v=>v.toFixed(2)));
    Plotly.react('heatmap',[{{z:M,x:DL,y:T,type:'heatmap',colorscale:cs,text:txt,texttemplate:'%{{text}}',
    textfont:{{size:10,color:'black'}},colorbar:{{title:{{text:barTitle,font:{{color:'#e0e0e0'}}}},
    tickfont:{{color:'#e0e0e0'}},len:.9}},zmin:zmin,zmax:zmax}}],
    {{margin:{{t:20,b:80,l:60,r:50}},xaxis:{{title:'Delta',tickangle:45,color:'#e0e0e0'}},
    yaxis:{{title:'Tenor',color:'#e0e0e0'}},paper_bgcolor:'#3d3d3d',plot_bgcolor:'#3d3d3d'}});
}}

function renderTables(){{
    var data=greekData.gamma; var n=getNotional();
    var cheapHtml='<table><tr><th>Tenor</th><th>Delta</th><th>Richness</th><th>Theta ($K)</th><th>Gamma ($M)</th></tr>';
    data.cheap.forEach(r=>{{
        var theta=Math.abs(r.theta*n),gamma=Math.abs(r.gamma*n);
        cheapHtml+='<tr class="cheap"><td>'+r.tenor+'</td><td>'+r.delta+'</td><td>'+r.score.toFixed(2)+'</td><td>'+theta.toFixed(2)+'</td><td>'+gamma.toFixed(4)+'</td></tr>';
    }});
    cheapHtml+='</table>';
    document.getElementById('cheap-table').innerHTML=cheapHtml;

    var richHtml='<table><tr><th>Tenor</th><th>Delta</th><th>Richness</th><th>Theta ($K)</th><th>Gamma ($M)</th></tr>';
    data.rich.forEach(r=>{{
        var theta=Math.abs(r.theta*n),gamma=Math.abs(r.gamma*n);
        richHtml+='<tr class="rich"><td>'+r.tenor+'</td><td>'+r.delta+'</td><td>'+r.score.toFixed(2)+'</td><td>'+theta.toFixed(2)+'</td><td>'+gamma.toFixed(4)+'</td></tr>';
    }});
    richHtml+='</table>';
    document.getElementById('rich-table').innerHTML=richHtml;
}}

function init(){{
    var st=document.getElementById('calc-t'); T.forEach(t=>st.innerHTML+='<option>'+t+'</option>');
    var sd=document.getElementById('calc-d'); D.forEach((d,i)=>sd.innerHTML+='<option value="'+d+'">'+DL[i]+'</option>');
    renderTables(); renderHeatmap();
}}

function getP(t,d){{return S.find(x=>x.tenor===t&&x.delta_label===d);}}

function calcGreeks(){{
    var p=getP(document.getElementById('calc-t').value,document.getElementById('calc-d').value);
    if(!p) return;
    var n=getNotional();
    var gamma=Math.abs(p.gamma*n),theta=Math.abs(p.theta*n),vega=Math.abs(p.vega*n),delta=Math.abs(p.delta);
    document.getElementById('calc-result').innerHTML='<div class="greeks-result">'+
    '<div style="margin-bottom:10px"><strong>'+p.tenor+' '+p.delta_label+'</strong> | K: '+p.strike.toFixed(4)+' | Vol: '+p.vol.toFixed(2)+'% | Days: '+p.days+'</div>'+
    '<div class="greeks-grid">'+
    '<div class="greek-box"><div class="greek-label">Delta</div><div class="greek-value">'+delta.toFixed(4)+'</div></div>'+
    '<div class="greek-box"><div class="greek-label">Gamma <span class="unit">($M)</span></div><div class="greek-value">'+gamma.toFixed(4)+'</div></div>'+
    '<div class="greek-box"><div class="greek-label">Theta/Day <span class="unit">($K)</span></div><div class="greek-value">'+theta.toFixed(2)+'</div></div>'+
    '<div class="greek-box"><div class="greek-label">Richness</div><div class="greek-value"><span class="richness-badge" style="background:'+richnessColor(p.richness)+'">'+p.richness.toFixed(2)+'</span></div></div>'+
    '</div>'+
    '<div style="margin-top:12px;padding:10px;background:#2a3f4f;border-radius:6px;border-left:4px solid #90caf9">'+
    '<div class="greeks-grid" style="grid-template-columns:1fr 1fr">'+
    '<div class="greek-box"><div class="greek-label">Vega <span class="unit">($K)</span></div><div class="greek-value">'+vega.toFixed(2)+'</div></div>'+
    '<div class="greek-box"><div class="greek-label">Richness Score</div><div class="greek-value"><span class="richness-badge" style="background:'+richnessColor(p.richness)+'">'+p.richness.toFixed(2)+'</span></div></div>'+
    '</div></div></div>';
}}

// === PORTFOLIO ANALYSIS ===

function normcdf(x){{
    var a1=0.254829592,a2=-0.284496736,a3=1.421413741,a4=-1.453152027,a5=1.061405429,p=0.3275911;
    var sign=x<0?-1:1; x=Math.abs(x)/Math.sqrt(2);
    var t=1/(1+p*x);
    var y=1-((((a5*t+a4)*t+a3)*t+a2)*t+a1)*t*Math.exp(-x*x);
    return 0.5*(1+sign*y);
}}
function normpdf(x){{return Math.exp(-0.5*x*x)/Math.sqrt(2*Math.PI);}}

// Interpolate forward and implied r_f for a given T from the forward curve
function interpForward(T){{
    if(fwdCurve.length===0) return {{F:mkt.spot,r_f:mkt.r_f}};
    // Clamp to curve boundaries
    if(T<=fwdCurve[0].T_years) return {{F:fwdCurve[0].forward,r_f:fwdCurve[0].r_f_implied}};
    if(T>=fwdCurve[fwdCurve.length-1].T_years) return {{F:fwdCurve[fwdCurve.length-1].forward,r_f:fwdCurve[fwdCurve.length-1].r_f_implied}};
    // Linear interpolation
    for(var i=0;i<fwdCurve.length-1;i++){{
        if(T>=fwdCurve[i].T_years&&T<=fwdCurve[i+1].T_years){{
            var t0=fwdCurve[i].T_years,t1=fwdCurve[i+1].T_years;
            var w=(T-t0)/(t1-t0);
            var F=fwdCurve[i].forward+(fwdCurve[i+1].forward-fwdCurve[i].forward)*w;
            var rf=fwdCurve[i].r_f_implied+(fwdCurve[i+1].r_f_implied-fwdCurve[i].r_f_implied)*w;
            return {{F:F,r_f:rf}};
        }}
    }}
    return {{F:mkt.spot,r_f:mkt.r_f}};
}}

function calcPositionGreeks(K,T,vol,notional,isCall){{
    var Sp=mkt.spot,rd=mkt.r_d,sigma=vol/100;
    if(T<=0) T=1/365;
    // Use interpolated forward to get correct r_f for this tenor
    var fwd=interpForward(T);
    var rf=fwd.r_f;
    var d1v=(Math.log(Sp/K)+(rd-rf+0.5*sigma*sigma)*T)/(sigma*Math.sqrt(T));
    var d2v=d1v-sigma*Math.sqrt(T);
    var df=Math.exp(-rd*T);
    var delta_raw=isCall?Math.exp(-rf*T)*normcdf(d1v):-Math.exp(-rf*T)*normcdf(-d1v);
    var gamma_raw=Math.exp(-rf*T)*normpdf(d1v)/(Sp*sigma*Math.sqrt(T));

    // Raw gamma cost formula: |theta + delta * tn_roll| / gamma (per-unit, no notional)
    var theta_per_unit=0;
    var t1=-Sp*normpdf(d1v)*sigma*Math.exp(-rf*T)/(2*Math.sqrt(T));
    var t2,t3;
    if(isCall){{t2=-rd*K*df*normcdf(d2v);t3=rf*Sp*Math.exp(-rf*T)*normcdf(d1v);}}
    else{{t2=rd*K*df*normcdf(-d2v);t3=-rf*Sp*Math.exp(-rf*T)*normcdf(-d1v);}}
    theta_per_unit=(t1+t2+t3)/365;

    var roll_per_unit=delta_raw*(mkt.tn_roll/10000);
    var gamma_cost_raw=gamma_raw>1e-12?Math.abs(theta_per_unit+roll_per_unit)/gamma_raw:0;
    var richness=rawToRichness(gamma_cost_raw);

    // Scaled Greeks for display
    var gamma_scaled=gamma_raw*Sp*0.01*Math.abs(notional);
    var theta_scaled=(theta_per_unit)*Math.abs(notional)*1000;
    var vega_raw=Sp*Math.exp(-rf*T)*normpdf(d1v)*Math.sqrt(T);
    var vega_scaled=vega_raw*0.01*Math.abs(notional)*1000;
    var days=Math.max(1,Math.round(T*365));
    var totalDecay=theta_scaled*days;
    var sign=notional>=0?1:-1;

    return{{
        delta_raw:delta_raw,
        delta:delta_raw*sign,
        gamma:gamma_scaled*sign,
        theta:theta_scaled*sign,
        vega:vega_scaled*sign,
        days:days,
        totalDecay:totalDecay*sign,
        gamma_cost_raw:gamma_cost_raw,
        richness:richness
    }};
}}

function getTenorBucket(days){{
    var buckets=[['O/N',1],['1W',7],['2W',14],['1M',30],['2M',61],['3M',91],['6M',182],['9M',274],['1Y',365],['2Y',730]];
    for(var i=buckets.length-1;i>=0;i--){{if(days>=buckets[i][1]*0.7) return buckets[i][0];}}
    return 'O/N';
}}

function getDeltaBucket(delta){{
    var d=Math.abs(delta)*100;
    var buckets=[10,15,20,25,30,35,40,45,50];
    var closest=50,minDiff=100;
    buckets.forEach(b=>{{if(Math.abs(d-b)<minDiff){{minDiff=Math.abs(d-b);closest=b;}}}});
    if(delta>=0) return closest===50?'ATM':closest+'C';
    else return closest===50?'ATM':closest+'P';
}}

function initPortfolio(){{
    var positions=[];
    excelPositions.forEach((p,i)=>{{
        var isCall=p.type==='C';
        var greeks=calcPositionGreeks(p.strike,p.T,p.vol,p.notional,isCall);
        var pos={{
            id:i+1,strike:p.strike,expiry:p.expiry,vol:p.vol,notional:p.notional,type:p.type,
            T:p.T,days:p.days,delta_raw:greeks.delta_raw,delta:greeks.delta,
            gamma:greeks.gamma,theta:greeks.theta,vega:greeks.vega,
            totalDecay:greeks.totalDecay,
            gamma_cost_raw:greeks.gamma_cost_raw,
            richness:greeks.richness,
            tenorBucket:getTenorBucket(p.days),deltaBucket:getDeltaBucket(greeks.delta)
        }};
        positions.push(pos);
    }});
    portfolioData={{positions:positions}};

    if(positions.length===0){{
        document.getElementById('portfolio-content').innerHTML='<div class="no-positions"><h2>No Positions Found</h2><p>Add positions to the "Positions" sheet in fx_gamma_inputs.xlsx and re-run the script.</p></div>';
        return;
    }}

    // Compute portfolio summary
    var totals={{gamma:0,theta:0,vega:0}};
    var richSum=0,gammaAbsSum=0;
    positions.forEach(p=>{{
        totals.gamma+=p.gamma; totals.theta+=p.theta; totals.vega+=p.vega;
        var gAbs=Math.abs(p.gamma);
        richSum+=p.richness*gAbs;
        gammaAbsSum+=gAbs;
    }});
    var wtdRichness=gammaAbsSum>0?(richSum/gammaAbsSum):0;

    document.getElementById('sum-gamma').textContent=totals.gamma.toFixed(4);
    var thetaLabel=(totals.theta<0?'Paying $':'Receiving $')+Math.abs(totals.theta).toFixed(2)+'K';
    document.getElementById('sum-theta').textContent=thetaLabel;
    document.getElementById('sum-theta').style.color=totals.theta<0?'#ef5350':'#66bb6a';
    document.getElementById('sum-theta').style.fontSize='14px';
    document.getElementById('sum-richness').textContent=wtdRichness.toFixed(2);
    document.getElementById('sum-richness').style.background=richnessColor(wtdRichness);
    document.getElementById('sum-richness').style.color='#000';
    document.getElementById('sum-richness').style.borderRadius='4px';
    document.getElementById('sum-richness').style.padding='2px 6px';
    document.getElementById('sum-vega').textContent=totals.vega.toFixed(2);

    // Position details table
    var html='<table><tr><th>#</th><th>Strike</th><th>Expiry</th><th>Days</th><th>Vol</th><th>Notl</th><th>Type</th><th>Delta</th><th>Gamma</th><th>Theta</th><th>Vega</th><th>Richness</th></tr>';
    positions.forEach(p=>{{
        var cls=p.notional>=0?'cheap':'rich';
        html+='<tr class="'+cls+'"><td>'+p.id+'</td><td>'+p.strike.toFixed(3)+'</td><td>'+p.expiry+'</td><td>'+p.days+'</td><td>'+p.vol.toFixed(1)+'</td><td>'+(p.notional>=0?'+':'')+p.notional.toFixed(1)+'</td><td>'+p.type+'</td><td>'+p.delta.toFixed(2)+'</td><td>'+p.gamma.toFixed(4)+'</td><td>'+p.theta.toFixed(2)+'</td><td>'+p.vega.toFixed(1)+'</td><td><span class="richness-badge" style="background:'+richnessColor(p.richness)+'">'+p.richness.toFixed(2)+'</span></td></tr>';
    }});
    html+='</table>';
    document.getElementById('positions-table').innerHTML=html;
    portfolioData.allPositions=positions;
    renderIneffTable();
    buildPortfolioHeatmap(positions);
    buildTimeSeries(positions);
}}

var currentIneffView='long';
function setIneffView(view){{
    currentIneffView=view;
    ['long','short'].forEach(v=>{{document.getElementById('ineff-btn-'+v).className='btn-toggle'+(v===view?' active':'');}});
    renderIneffTable();
}}

function renderIneffTable(){{
    if(!portfolioData||!portfolioData.allPositions) return;
    var positions=portfolioData.allPositions;
    var filtered,sorted;
    if(currentIneffView==='long'){{
        // Longs with highest richness = overpaying for gamma
        filtered=positions.filter(p=>p.notional>0);
        sorted=filtered.sort((a,b)=>b.richness-a.richness);
    }}else{{
        // Shorts with lowest richness = selling cheap gamma (sorted high to low)
        filtered=positions.filter(p=>p.notional<0);
        sorted=filtered.sort((a,b)=>b.richness-a.richness);
    }}
    if(sorted.length===0){{document.getElementById('ineff-table').innerHTML='<p style="text-align:center;color:#aaa">No '+(currentIneffView==='long'?'long':'short')+' positions</p>';return;}}
    var html='<table><tr><th>#</th><th>Strike</th><th>Expiry</th><th>Days</th><th>Notl</th><th>Type</th><th>Theta</th><th>Total Decay</th><th>Richness</th></tr>';
    sorted.forEach(p=>{{
        var cls=p.notional>=0?'cheap':'rich';
        html+='<tr class="'+cls+'"><td>'+p.id+'</td><td>'+p.strike.toFixed(3)+'</td><td>'+p.expiry+'</td><td>'+p.days+'</td><td>'+(p.notional>=0?'+':'')+p.notional.toFixed(1)+'</td><td>'+p.type+'</td><td>'+p.theta.toFixed(2)+'</td><td>'+p.totalDecay.toFixed(1)+'</td><td><span class="richness-badge" style="background:'+richnessColor(p.richness)+'">'+p.richness.toFixed(2)+'</span></td></tr>';
    }});
    html+='</table>';
    document.getElementById('ineff-table').innerHTML=html;
}}

function buildPortfolioHeatmap(positions){{
    var tenorList=T;
    var allStrikes=positions.map(p=>p.strike);
    var minS=Math.min(...allStrikes),maxS=Math.max(...allStrikes),range=maxS-minS;
    var maxBuckets=18,rawWidth=range/maxBuckets;
    var niceWidths=[0.001,0.0025,0.005,0.01,0.025,0.05,0.1,0.25,0.5];
    var bucketWidth=niceWidths.find(w=>w>=rawWidth)||0.5;
    var uniqueStrikes=[...new Set(allStrikes)].sort((a,b)=>a-b);
    var useBuckets=uniqueStrikes.length>maxBuckets;
    var strikeBuckets,bucketLabels;
    if(useBuckets){{
        var bStart=Math.floor(minS/bucketWidth)*bucketWidth,bEnd=Math.ceil(maxS/bucketWidth)*bucketWidth;
        strikeBuckets=[]; bucketLabels=[];
        for(var b=bStart;b<bEnd+bucketWidth;b+=bucketWidth){{strikeBuckets.push(b);bucketLabels.push(b.toFixed(3));}}
    }}else{{strikeBuckets=uniqueStrikes;bucketLabels=uniqueStrikes.map(s=>s.toFixed(3));}}
    portfolioData.strikes=strikeBuckets;portfolioData.strikeLabels=bucketLabels;portfolioData.bucketWidth=useBuckets?bucketWidth:0;
    var nb=strikeBuckets.length;
    portfolioData.heatmaps={{
        // Gamma-weighted average richness per bucket
        richness:tenorList.map(()=>strikeBuckets.map(()=>null)),
        vega:tenorList.map(()=>strikeBuckets.map(()=>0)),
        positionsByBucket:tenorList.map(()=>strikeBuckets.map(()=>[]))
    }};
    function getSI(strike){{
        if(!useBuckets) return strikeBuckets.indexOf(strike);
        return Math.max(0,Math.min(Math.floor((strike-strikeBuckets[0])/bucketWidth),nb-1));
    }}
    // Accumulate positions into buckets
    var bucketGammaSum=tenorList.map(()=>strikeBuckets.map(()=>0));
    var bucketRichGammaSum=tenorList.map(()=>strikeBuckets.map(()=>0));
    var bucketNetNotional=tenorList.map(()=>strikeBuckets.map(()=>0));
    positions.forEach(p=>{{
        var ti=tenorList.indexOf(p.tenorBucket),si=getSI(p.strike);
        if(ti>=0&&si>=0){{
            portfolioData.heatmaps.vega[ti][si]+=p.vega;
            portfolioData.heatmaps.positionsByBucket[ti][si].push(p);
            var gAbs=Math.abs(p.gamma);
            bucketGammaSum[ti][si]+=gAbs;
            bucketRichGammaSum[ti][si]+=p.richness*gAbs;
            bucketNetNotional[ti][si]+=p.notional;
        }}
    }});
    // Compute gamma-weighted average richness per bucket
    portfolioData.heatmaps.direction=tenorList.map(()=>strikeBuckets.map(()=>null));
    for(var ti=0;ti<tenorList.length;ti++){{
        for(var si=0;si<nb;si++){{
            if(bucketGammaSum[ti][si]>1e-10){{
                portfolioData.heatmaps.richness[ti][si]=bucketRichGammaSum[ti][si]/bucketGammaSum[ti][si];
                portfolioData.heatmaps.direction[ti][si]=bucketNetNotional[ti][si]>=0?'L':'S';
            }}
        }}
    }}
}}

function setPortHeatmapView(view){{
    currentPortHeatmap=view;
    ['richness','vega'].forEach(v=>{{document.getElementById('port-btn-'+v).className='btn-toggle'+(v===view?' active':'');}});
    var titles={{richness:'Portfolio Richness (1-5)',vega:'Portfolio Vega'}};
    document.getElementById('port-heatmap-title').textContent=titles[view];
    renderPortHeatmap();
}}

function renderPortHeatmap(){{
    if(!portfolioData||!portfolioData.heatmaps||!portfolioData.strikeLabels) return;
    var M,cs,zmin,zmax;
    if(currentPortHeatmap==='richness'){{
        M=portfolioData.heatmaps.richness;
        cs=cs_richness; zmin=1; zmax=5;
    }}else{{
        M=portfolioData.heatmaps.vega.map(r=>r.map(v=>v===0?null:v));
        cs=[[0,'#f44336'],[0.5,'#9e9e9e'],[1,'#4caf50']];
        var flat=M.flat().filter(v=>v!==null);
        var mx=flat.length>0?Math.max(Math.abs(Math.min(...flat)),Math.abs(Math.max(...flat))):1;
        zmin=-mx; zmax=mx;
    }}
    var xLabels=portfolioData.strikeLabels;
    var dir=portfolioData.heatmaps.direction;
    var txt;
    if(currentPortHeatmap==='richness'){{
        txt=M.map((r,ti)=>r.map((v,si)=>v===null?'':(dir[ti][si]||'')+' '+v.toFixed(2)));
    }}else{{
        txt=M.map((r,ti)=>r.map((v,si)=>{{
            if(v===null) return '';
            var d=dir[ti]&&dir[ti][si]?dir[ti][si]:'';
            return d+' '+v.toFixed(1);
        }}));
    }}
    Plotly.react('port-heatmap',[{{z:M,x:xLabels,y:T,type:'heatmap',colorscale:cs,text:txt,texttemplate:'%{{text}}',
    textfont:{{size:9,color:'black'}},colorbar:{{title:{{text:currentPortHeatmap==='richness'?'Score':'$K',font:{{color:'#e0e0e0'}}}},tickfont:{{color:'#e0e0e0'}},len:.9}},
    zmin:zmin,zmax:zmax,hoverongaps:false}}],
    {{margin:{{t:20,b:80,l:60,r:50}},xaxis:{{title:'Strike',tickangle:45,color:'#e0e0e0',type:'category'}},
    yaxis:{{title:'Tenor',color:'#e0e0e0'}},paper_bgcolor:'#3d3d3d',plot_bgcolor:'#3d3d3d'}});
    var hm=document.getElementById('port-heatmap');
    hm.removeAllListeners&&hm.removeAllListeners('plotly_click');
    hm.on('plotly_click',function(data){{
        var pt=data.points[0],ti=T.indexOf(pt.y),si=pt.pointIndex[1];
        if(ti>=0&&si>=0) showDrillDown(ti,si);
    }});
}}

function showDrillDown(ti,si){{
    if(!portfolioData||!portfolioData.heatmaps.positionsByBucket) return;
    var positions=portfolioData.heatmaps.positionsByBucket[ti][si];
    var tenor=T[ti],strike=portfolioData.strikes[si],bw=portfolioData.bucketWidth||0;
    var strikeLabel=bw>0?strike.toFixed(3)+' - '+(strike+bw).toFixed(3):strike.toFixed(3);
    document.getElementById('drill-title').textContent=tenor+' / Strike '+strikeLabel+' ('+positions.length+' positions)';
    if(positions.length===0){{
        document.getElementById('drill-content').innerHTML='<p style="text-align:center;color:#aaa">No positions in this bucket</p>';
    }}else{{
        // Compute bucket-level gamma-weighted richness
        var gAbsSum=0,richGammaSum=0,totGamma=0,totTheta=0,totVega=0;
        positions.forEach(p=>{{
            var gAbs=Math.abs(p.gamma);
            gAbsSum+=gAbs; richGammaSum+=p.richness*gAbs;
            totGamma+=p.gamma; totTheta+=p.theta; totVega+=p.vega;
        }});
        var bucketRichness=gAbsSum>0?(richGammaSum/gAbsSum):0;

        var html='<div style="margin-bottom:15px;padding:10px;background:#2d2d2d;border-radius:6px">'+
            '<strong>Bucket Summary:</strong> '+
            'Wtd Richness = <span class="richness-badge" style="background:'+richnessColor(bucketRichness)+'">'+bucketRichness.toFixed(2)+'</span> | '+
            '\u03b3 = '+totGamma.toFixed(4)+' | \u03b8 = '+totTheta.toFixed(2)+' | \u03bd = '+totVega.toFixed(1)+
            '</div>';

        html+='<table><tr><th>#</th><th>Strike</th><th>Expiry</th><th>Vol</th><th>Notl</th><th>Type</th><th>Gamma</th><th>Theta</th><th>Vega</th><th>Richness</th></tr>';
        positions.forEach(p=>{{
            var cls=p.notional>=0?'cheap':'rich';
            html+='<tr class="'+cls+'"><td>'+p.id+'</td><td>'+p.strike.toFixed(3)+'</td><td>'+p.expiry+'</td><td>'+p.vol.toFixed(1)+'</td><td>'+(p.notional>=0?'+':'')+p.notional.toFixed(1)+'</td><td>'+p.type+'</td><td>'+p.gamma.toFixed(4)+'</td><td>'+p.theta.toFixed(2)+'</td><td>'+p.vega.toFixed(1)+'</td><td><span class="richness-badge" style="background:'+richnessColor(p.richness)+'">'+p.richness.toFixed(2)+'</span></td></tr>';
        }});
        html+='</table>';
        document.getElementById('drill-content').innerHTML=html;
    }}
    document.getElementById('drill-modal').classList.add('show');
}}

function closeDrillDown(){{
    document.getElementById('drill-modal').classList.remove('show');
}}

// === PORTFOLIO GREEKS TIME SERIES ===

var currentTimeView='gamma';
var timeSeriesData=null;

function setTimeView(view){{
    currentTimeView=view;
    ['gamma','theta','vega','cumDecay'].forEach(v=>{{
        document.getElementById('time-btn-'+v).className='btn-toggle'+(v===view?' active':'');
    }});
    renderTimeSeries();
}}

function buildTimeSeries(positions){{
    if(!positions||positions.length===0){{timeSeriesData=null;return;}}
    // Find max days across all positions
    var maxDays=Math.max(...positions.map(p=>p.days));
    // Cap at 365 for readability; if longer positions exist, still project them
    var projDays=Math.min(maxDays,730);
    var dates=[],gammas=[],thetas=[],vegas=[],deltas=[],cumDecays=[];
    var cumDecay=0;
    var Sp=mkt.spot,rd=mkt.r_d;

    for(var day=0;day<=projDays;day++){{
        var dStr=new Date(Date.now()+day*86400000).toISOString().slice(0,10);
        dates.push(dStr);
        var totGamma=0,totTheta=0,totVega=0,totDelta=0;
        positions.forEach(p=>{{
            var remainT=p.T-day/365;
            if(remainT<1/365) return; // expired
            var sigma=p.vol/100,K=p.strike,isCall=p.type==='C';
            var fwd=interpForward(remainT);
            var rf=fwd.r_f;
            var d1v=(Math.log(Sp/K)+(rd-rf+0.5*sigma*sigma)*remainT)/(sigma*Math.sqrt(remainT));
            var d2v=d1v-sigma*Math.sqrt(remainT);
            var dfr=Math.exp(-rd*remainT),dff=Math.exp(-rf*remainT);
            // Delta
            var delta_raw=isCall?dff*normcdf(d1v):-dff*normcdf(-d1v);
            // Gamma
            var gamma_raw=dff*normpdf(d1v)/(Sp*sigma*Math.sqrt(remainT));
            var gamma_scaled=gamma_raw*Sp*0.01*Math.abs(p.notional);
            // Theta
            var t1=-Sp*normpdf(d1v)*sigma*dff/(2*Math.sqrt(remainT));
            var t2,t3;
            if(isCall){{t2=-rd*K*dfr*normcdf(d2v);t3=rf*Sp*dff*normcdf(d1v);}}
            else{{t2=rd*K*dfr*normcdf(-d2v);t3=-rf*Sp*dff*normcdf(-d1v);}}
            var theta_daily=((t1+t2+t3)/365)*Math.abs(p.notional)*1000;
            // Vega
            var vega_raw=Sp*dff*normpdf(d1v)*Math.sqrt(remainT);
            var vega_scaled=vega_raw*0.01*Math.abs(p.notional)*1000;
            var sign=p.notional>=0?1:-1;
            totGamma+=gamma_scaled*sign;
            totTheta+=theta_daily*sign;
            totVega+=vega_scaled*sign;
            totDelta+=delta_raw*sign;
        }});
        gammas.push(totGamma);
        thetas.push(totTheta);
        vegas.push(totVega);
        deltas.push(totDelta);
        cumDecay+=totTheta;
        cumDecays.push(cumDecay);
    }}
    timeSeriesData={{dates:dates,gamma:gammas,theta:thetas,vega:vegas,delta:deltas,cumDecay:cumDecays}};
    // Update summary with actual projected cumulative decay
    if(cumDecays.length>0){{
        var finalDecay=cumDecays[cumDecays.length-1];
        var decayAbs=Math.abs(finalDecay);
        var decayStr=decayAbs>=1000?(decayAbs/1000).toFixed(2)+'M':decayAbs.toFixed(1)+'K';
        var decayLabel=(finalDecay<0?'Paying $':'Receiving $')+decayStr;
        document.getElementById('sum-decay').textContent=decayLabel;
        document.getElementById('sum-decay').style.color=finalDecay<0?'#ef5350':'#66bb6a';
        document.getElementById('sum-decay').style.fontSize='14px';
    }}
    renderTimeSeries();
}}

function renderTimeSeries(){{
    if(!timeSeriesData) return;
    var d=timeSeriesData;
    var viewMap={{gamma:{{y:d.gamma,name:'Gamma ($M)',color:'#64b5f6'}},
                  theta:{{y:d.theta,name:'Theta ($K/day)',color:'#ef5350'}},
                  vega:{{y:d.vega,name:'Vega ($K)',color:'#ab47bc'}},
                  delta:{{y:d.delta,name:'Delta',color:'#66bb6a'}},
                  cumDecay:{{y:d.cumDecay,name:'Cumulative Decay ($K)',color:'#ffa726'}}}};
    var v=viewMap[currentTimeView];
    // Fill color: positive area green-ish, negative area red-ish
    var hasNeg=v.y.some(val=>val<0),hasPos=v.y.some(val=>val>0);
    var traces=[];
    if(currentTimeView==='cumDecay'||currentTimeView==='theta'){{
        // Show zero line prominently, fill to zero
        traces.push({{x:d.dates,y:v.y,type:'scatter',mode:'lines',name:v.name,
            line:{{color:v.color,width:2}},fill:'tozeroy',
            fillcolor:hasNeg?'rgba(239,83,80,0.15)':'rgba(102,187,106,0.15)'}});
    }}else{{
        traces.push({{x:d.dates,y:v.y,type:'scatter',mode:'lines',name:v.name,
            line:{{color:v.color,width:2}}}});
    }}
    var yTitle={{gamma:'$M',theta:'$K/day',vega:'$K',delta:'Delta',cumDecay:'$K'}};
    Plotly.react('time-chart',traces,{{
        margin:{{t:20,b:50,l:60,r:30}},
        xaxis:{{title:'Date',color:'#e0e0e0',gridcolor:'#555',type:'date'}},
        yaxis:{{title:yTitle[currentTimeView],color:'#e0e0e0',gridcolor:'#555',zeroline:true,zerolinecolor:'#888',zerolinewidth:2}},
        paper_bgcolor:'#3d3d3d',plot_bgcolor:'#3d3d3d',
        showlegend:false,
        hovermode:'x unified',
        hoverlabel:{{font:{{color:'#ffffff'}},bgcolor:'#3d3d3d'}}
    }});
}}

init();
</script></body></html>'''

    with open(output, 'w', encoding='utf-8') as f: f.write(html)
    return output

# === MAIN ===

def main():
    fp = 'fx_gamma_inputs.xlsx'
    print("\n" + "="*50 + "\n   FX GAMMA TRADING DASHBOARD\n" + "="*50 + "\n")

    if not os.path.exists(fp):
        print(f"Creating template: {fp}")
        create_template(fp)
        print("\n  -> Edit Excel file, then run again.\n")
        return

    print(f"Loading: {fp}")
    mkt = load_data(fp)
    print(f"  Spot: {mkt['spot']:.4f} | Terms: {mkt['r_terms']*100:.2f}% | Base: {mkt['r_base']*100:.2f}%")

    print("\nBuilding surface...")
    df, fwd_df, p5, p95 = build_surface(mkt)
    on_fwd = fwd_df.loc[fwd_df['T_years'].idxmin(), 'fwd_points'] if len(fwd_df) > 0 else 0
    print(f"  O/N Fwd Pts: {on_fwd:.2f} pips | Normalization: p5={p5:.6f}, p95={p95:.6f}")

    print("\nForward Points:")
    print(fwd_df[['tenor', 'days', 'forward', 'fwd_points']].to_string(index=False))

    print("\nGamma Richness (1=Cheap, 5=Rich):")
    piv = df.pivot(index='tenor', columns='delta_label', values='richness')[['10P','25P','ATM','25C','10C']]
    print(piv.round(2).to_string())

    out = create_dashboard(df, fwd_df, mkt, p5, p95)
    print(f"\n  -> Dashboard: {out}")
    df.to_csv('fx_gamma_surface.csv', index=False)
    print(f"  -> CSV: fx_gamma_surface.csv\n\nComplete!\n")

if __name__ == '__main__':
    main()
