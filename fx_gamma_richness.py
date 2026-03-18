#!/usr/bin/env python3
"""
FX Options Analytics v2
- Market Data tab: mark vol surfaces per pair, set spot/rates/fwd pts
- Surface Analysis: gamma richness from your marks (computed live in JS)
- Portfolio Analysis: positions from Excel, Greeks against your marked surface
- DTCC Live Surface: real-time shifts on your marks from DTCC prints

Excel is ONLY for uploading positions (with pair column).
"""
import numpy as np, pandas as pd, json, os

def create_template(fp):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    wb=Workbook(); hf,hfill=Font(bold=True,color='FFFFFF'),PatternFill('solid',fgColor='1F4E79')
    inf,infill=Font(color='0000FF'),PatternFill('solid',fgColor='D6EAF8')
    bord=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ws=wb.active; ws.title='Positions'
    for i,h in enumerate(['Pair','Strike','Expiry','Vol','Notional','Type'],1):
        c=ws.cell(row=1,column=i,value=h); c.font,c.fill,c.border=hf,hfill,bord
    notes=['# Pair: EURUSD, USDJPY, etc','# Strike: option strike price','# Expiry: YYYY-MM-DD','# Vol: implied vol %','# Notional: $M (neg=short)','# Type: C or P']
    for i,n in enumerate(notes,2): ws.cell(row=i,column=1,value=n).font=Font(italic=True,color='888888')
    today=datetime.now()
    sample=[('EURUSD',1.09,(today+timedelta(90)).strftime('%Y-%m-%d'),8.5,10,'C'),
            ('EURUSD',1.075,(today+timedelta(45)).strftime('%Y-%m-%d'),9.2,-5,'P'),
            ('EURUSD',1.10,(today+timedelta(180)).strftime('%Y-%m-%d'),7.8,15,'C'),
            ('EURUSD',1.08,(today+timedelta(60)).strftime('%Y-%m-%d'),8.8,-8,'C'),
            ('USDJPY',155.0,(today+timedelta(30)).strftime('%Y-%m-%d'),9.0,20,'P'),
            ('USDJPY',152.0,(today+timedelta(90)).strftime('%Y-%m-%d'),8.5,-10,'C'),
            ('GBPUSD',1.27,(today+timedelta(60)).strftime('%Y-%m-%d'),7.2,12,'C')]
    for ri,row in enumerate(sample,9):
        for ci,v in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=v); c.border=bord; c.font,c.fill=inf,infill
    for i,w in enumerate([12,12,12,8,12,8],1): ws.column_dimensions[get_column_letter(i)].width=w
    wb.save(fp)

def load_positions(fp):
    from datetime import datetime
    df=pd.read_excel(fp,sheet_name='Positions')
    df.columns=['pair','strike','expiry','vol','notional','type'][:df.shape[1]]
    df=df[df['strike'].apply(lambda x:isinstance(x,(int,float)) and not pd.isna(x))]
    positions=[]; today=datetime.now()
    for _,row in df.iterrows():
        try:
            pair=str(row['pair']).upper().strip(); expiry=pd.to_datetime(row['expiry'])
            T=max(1/365,(expiry-pd.Timestamp(today)).days/365)
            positions.append({'pair':pair,'strike':float(row['strike']),'expiry':expiry.strftime('%Y-%m-%d'),
                'vol':float(row['vol']),'notional':float(row['notional']),'type':str(row['type']).upper().strip(),
                'T':round(T,6),'days':max(1,int(T*365))})
        except: continue
    return positions

DEFAULT_SURFACES = {
  'EURUSD':{'spot':1.085,'r_d':0.045,'r_f':0.025,'tenors':[
    {'tenor':'O/N','T':0.00274,'atm':7.5,'rr25':-0.3,'rr10':-0.6,'fly25':0.15,'fly10':0.4,'fwdPts':0.59},
    {'tenor':'1W','T':0.01918,'atm':7.8,'rr25':-0.35,'rr10':-0.7,'fly25':0.18,'fly10':0.45,'fwdPts':4.16},
    {'tenor':'2W','T':0.03836,'atm':8.0,'rr25':-0.4,'rr10':-0.8,'fly25':0.2,'fly10':0.5,'fwdPts':8.33},
    {'tenor':'1M','T':0.08333,'atm':8.2,'rr25':-0.5,'rr10':-1.0,'fly25':0.25,'fly10':0.6,'fwdPts':18.1},
    {'tenor':'2M','T':0.16667,'atm':8.5,'rr25':-0.6,'rr10':-1.2,'fly25':0.3,'fly10':0.7,'fwdPts':36.23},
    {'tenor':'3M','T':0.25,'atm':8.8,'rr25':-0.7,'rr10':-1.4,'fly25':0.35,'fly10':0.8,'fwdPts':54.39},
    {'tenor':'6M','T':0.5,'atm':9.2,'rr25':-0.8,'rr10':-1.6,'fly25':0.4,'fly10':0.95,'fwdPts':109.04},
    {'tenor':'9M','T':0.75,'atm':9.5,'rr25':-0.85,'rr10':-1.7,'fly25':0.42,'fly10':1.05,'fwdPts':163.98},
    {'tenor':'1Y','T':1.0,'atm':9.8,'rr25':-0.9,'rr10':-1.8,'fly25':0.45,'fly10':1.15,'fwdPts':219.18},
    {'tenor':'2Y','T':2.0,'atm':10.2,'rr25':-1.0,'rr10':-2.0,'fly25':0.5,'fly10':1.3,'fwdPts':442.8}]},
  'USDJPY':{'spot':149.5,'r_d':0.045,'r_f':0.005,'tenors':[
    {'tenor':'O/N','T':0.00274,'atm':9.5,'rr25':0.4,'rr10':1.1,'fly25':0.35,'fly10':1.0,'fwdPts':-0.45},
    {'tenor':'1W','T':0.01918,'atm':9.0,'rr25':0.35,'rr10':0.9,'fly25':0.3,'fly10':0.85,'fwdPts':-3.15},
    {'tenor':'2W','T':0.03836,'atm':8.8,'rr25':0.3,'rr10':0.8,'fly25':0.25,'fly10':0.75,'fwdPts':-6.3},
    {'tenor':'1M','T':0.08333,'atm':8.5,'rr25':0.25,'rr10':0.7,'fly25':0.2,'fly10':0.65,'fwdPts':-13.7},
    {'tenor':'2M','T':0.16667,'atm':8.3,'rr25':0.2,'rr10':0.6,'fly25':0.18,'fly10':0.55,'fwdPts':-27.4},
    {'tenor':'3M','T':0.25,'atm':8.2,'rr25':0.2,'rr10':0.55,'fly25':0.16,'fly10':0.5,'fwdPts':-41.1},
    {'tenor':'6M','T':0.5,'atm':8.4,'rr25':0.15,'rr10':0.5,'fly25':0.15,'fly10':0.45,'fwdPts':-82.2},
    {'tenor':'9M','T':0.75,'atm':8.6,'rr25':0.15,'rr10':0.45,'fly25':0.14,'fly10':0.42,'fwdPts':-123.3},
    {'tenor':'1Y','T':1.0,'atm':8.8,'rr25':0.1,'rr10':0.4,'fly25':0.13,'fly10':0.4,'fwdPts':-164.4},
    {'tenor':'2Y','T':2.0,'atm':9.2,'rr25':0.1,'rr10':0.35,'fly25':0.12,'fly10':0.38,'fwdPts':-328.8}]},
  'GBPUSD':{'spot':1.264,'r_d':0.045,'r_f':0.044,'tenors':[
    {'tenor':'O/N','T':0.00274,'atm':8.2,'rr25':-0.2,'rr10':-0.7,'fly25':0.3,'fly10':0.9,'fwdPts':0.01},
    {'tenor':'1W','T':0.01918,'atm':7.6,'rr25':-0.15,'rr10':-0.5,'fly25':0.25,'fly10':0.75,'fwdPts':0.07},
    {'tenor':'2W','T':0.03836,'atm':7.3,'rr25':-0.12,'rr10':-0.4,'fly25':0.2,'fly10':0.65,'fwdPts':0.14},
    {'tenor':'1M','T':0.08333,'atm':7.0,'rr25':-0.1,'rr10':-0.35,'fly25':0.18,'fly10':0.55,'fwdPts':0.3},
    {'tenor':'2M','T':0.16667,'atm':6.8,'rr25':-0.08,'rr10':-0.3,'fly25':0.15,'fly10':0.48,'fwdPts':0.6},
    {'tenor':'3M','T':0.25,'atm':6.9,'rr25':-0.08,'rr10':-0.28,'fly25':0.14,'fly10':0.43,'fwdPts':0.9},
    {'tenor':'6M','T':0.5,'atm':7.1,'rr25':-0.07,'rr10':-0.25,'fly25':0.13,'fly10':0.4,'fwdPts':1.8},
    {'tenor':'9M','T':0.75,'atm':7.3,'rr25':-0.07,'rr10':-0.23,'fly25':0.12,'fly10':0.37,'fwdPts':2.7},
    {'tenor':'1Y','T':1.0,'atm':7.5,'rr25':-0.06,'rr10':-0.2,'fly25':0.11,'fly10':0.35,'fwdPts':3.6},
    {'tenor':'2Y','T':2.0,'atm':7.9,'rr25':-0.05,'rr10':-0.18,'fly25':0.1,'fly10':0.33,'fwdPts':7.2}]}
}

CSS = '''*{box-sizing:border-box}body{font-family:'Segoe UI',Arial,sans-serif;margin:0;padding:15px;background:#2d2d2d;color:#e0e0e0}
h1{text-align:center;color:#90caf9;margin:10px 0 15px}h2{color:#90caf9;margin:0 0 15px;font-size:18px}h3{color:#90caf9;margin:15px 0 10px;font-size:14px}
.card{background:#3d3d3d;padding:15px;border-radius:8px;box-shadow:0 2px 4px rgba(0,0,0,0.3);margin-bottom:15px}
table{width:100%;border-collapse:collapse;font-size:11px;margin-bottom:10px}th,td{padding:5px 6px;text-align:right;border:1px solid #555}
th{background:#1F4E79;color:white}td:first-child{text-align:left;font-weight:600}
.tab-bar{display:flex;gap:0;margin-bottom:15px;max-width:1600px;margin-left:auto;margin-right:auto}
.tab-btn{padding:12px 24px;border:none;background:#3d3d3d;color:#aaa;font-size:14px;font-weight:600;cursor:pointer;border-bottom:3px solid transparent;transition:all 0.2s}
.tab-btn:hover{color:#e0e0e0;background:#454545}.tab-btn.active{color:#90caf9;border-bottom-color:#90caf9}
.tab-content{display:none;max-width:1600px;margin:0 auto}.tab-content.active{display:block}
.btn{padding:6px 14px;border:none;border-radius:4px;cursor:pointer;font-weight:600;font-size:12px}
.btn-primary{background:#1F4E79;color:white}.btn-toggle{padding:6px 12px;border:1px solid #555;border-radius:4px;cursor:pointer;font-weight:600;margin-right:4px;background:#2d2d2d;color:#e0e0e0;font-size:12px}
.btn-toggle.active{background:#1F4E79;color:white;border-color:#1F4E79}
.help{font-size:11px;color:#aaa;margin-top:8px}select{padding:6px;background:#3d3d3d;color:#e0e0e0;border:1px solid #555;border-radius:4px}
.grid{display:grid;grid-template-columns:1fr 420px;gap:15px}
.legend{display:flex;justify-content:center;gap:20px;margin:10px 0;font-size:12px}
.legend-item{display:flex;align-items:center;gap:5px}.legend-color{width:16px;height:16px;border-radius:3px}
.market-bar{display:flex;gap:20px;justify-content:center;margin-bottom:15px;padding:10px;background:#3d3d3d;border-radius:8px;flex-wrap:wrap}
.market-item{text-align:center}.market-item .label{font-size:11px;color:#aaa}.market-item .value{font-size:16px;font-weight:bold;color:#90caf9}
.notional-input{width:80px;padding:6px;background:#2d2d2d;color:#90caf9;border:1px solid #555;border-radius:4px;font-size:14px;font-weight:bold;text-align:center}
.cheap{background:#1a3a4a}.rich{background:#4a2a2a}
.portfolio-grid{display:grid;grid-template-columns:1fr 1fr;gap:15px}
.summary-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:15px}
.summary-box{background:#2d2d2d;padding:12px;border-radius:6px;text-align:center}
.summary-label{font-size:11px;color:#aaa;margin-bottom:4px}.summary-value{font-size:18px;font-weight:bold;color:#90caf9}
.richness-badge{display:inline-block;padding:2px 6px;border-radius:3px;font-weight:bold;font-size:11px;color:#000}
.pair-tab{padding:6px 14px;border:1px solid #555;border-radius:4px;cursor:pointer;font-size:12px;font-weight:600;background:#2d2d2d;color:#aaa;transition:all 0.15s}
.pair-tab:hover{border-color:#777;color:#e0e0e0}.pair-tab.active{background:#1F4E79;color:white;border-color:#1F4E79}
.add-pair-btn{padding:6px 12px;border:1px dashed #555;border-radius:4px;cursor:pointer;font-size:12px;color:#888;background:transparent}
.add-pair-btn:hover{border-color:#90caf9;color:#90caf9}
.vol-table{width:100%;border-collapse:collapse;font-size:12px}
.vol-table th{background:#1F4E79;color:white;padding:6px 8px;font-size:11px}
.vol-table td{padding:4px 6px;text-align:center}
.vol-table input{width:60px;padding:3px 5px;background:#2d2d2d;color:#e0e0e0;border:1px solid #555;border-radius:3px;font-size:12px;text-align:center}
.vol-table input:focus{border-color:#90caf9;outline:none;color:#90caf9}
.vol-table td:first-child{text-align:left;font-weight:600;color:#90caf9;background:#3a3a3a}
.mkt-param-row{display:flex;gap:15px;margin-bottom:12px;flex-wrap:wrap;align-items:center}
.mkt-param{display:flex;align-items:center;gap:6px}
.mkt-param label{font-size:11px;color:#aaa;white-space:nowrap}
.mkt-input{width:90px;padding:5px 8px;background:#2d2d2d;color:#90caf9;border:1px solid #555;border-radius:4px;font-size:13px;font-weight:600;text-align:center}
.mkt-input:focus{border-color:#90caf9;outline:none}
.dtcc-container{max-width:1600px;margin:0 auto}
.dtcc-pair-bar{display:flex;gap:4px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.dtcc-pair-label{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:#888;margin-right:4px}
.dtcc-pair-chip{padding:4px 10px;font-size:11px;font-weight:600;border-radius:4px;cursor:pointer;border:1px solid #555;background:#2d2d2d;color:#aaa}
.dtcc-pair-chip:hover{border-color:#777;color:#e0e0e0}.dtcc-pair-chip.active{background:#1F4E79;color:white;border-color:#1F4E79}
.dtcc-pair-divider{width:1px;height:20px;background:#555;margin:0 6px}
.dtcc-grid{display:grid;grid-template-columns:2fr 1fr;gap:15px}
.dtcc-strat-table{width:100%;border-collapse:collapse}
.dtcc-strat-table th{text-align:right;padding:5px 7px;font-size:10px;color:#888;text-transform:uppercase;border-bottom:1px solid #555}
.dtcc-strat-table th:first-child{text-align:left}
.dtcc-strat-table td{padding:5px 7px;text-align:right;font-size:12px;font-variant-numeric:tabular-nums;border-bottom:1px solid #444}
.dtcc-strat-table td:first-child{text-align:left;color:#aaa;font-weight:600}
.dtcc-chg-up{color:#66bb6a}.dtcc-chg-dn{color:#ef5350}.dtcc-chg-flat{color:#555}
.dtcc-feed-panel{max-height:500px;overflow-y:auto}
.dtcc-feed-item{padding:5px 8px;border-bottom:1px solid #444;font-size:11px;display:grid;grid-template-columns:55px 38px 45px 42px 50px 40px;gap:4px;align-items:center}
.dtcc-feed-item.fresh{background:rgba(144,202,249,0.08)}
.dtcc-feed-header{padding:4px 8px;font-size:9px;font-weight:700;text-transform:uppercase;color:#888;display:grid;grid-template-columns:55px 38px 45px 42px 50px 40px;gap:4px;border-bottom:1px solid #555;background:#2d2d2d;position:sticky;top:0;z-index:1}
.dtcc-dot{width:7px;height:7px;border-radius:50%;display:inline-block}
.dtcc-dot.ok{background:#66bb6a;animation:pulse 2s infinite}.dtcc-dot.err{background:#ef5350}.dtcc-dot.loading{background:#ffa726;animation:pulse 1s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.5}}
.dtcc-snap-badge{font-size:11px;color:#888;background:#2d2d2d;padding:2px 8px;border-radius:4px}
.dtcc-summary-row{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap}
.dtcc-summary-badge{background:#2d2d2d;border:1px solid #555;border-radius:6px;padding:8px 12px;min-width:130px}
.dtcc-summary-badge .dl{font-size:9px;color:#888;text-transform:uppercase}.dtcc-summary-badge .dv{font-size:18px;font-weight:700;margin-top:2px;color:#90caf9}.dtcc-summary-badge .ds{font-size:11px;color:#aaa}
.js-plotly-plot .plotly .modebar{display:none!important}'''

def create_dashboard(positions, output='fx_gamma_trading.html'):
    pos_json = json.dumps(positions)
    pairs_in_pos = json.dumps(list(set(p['pair'] for p in positions)))
    defaults_json = json.dumps(DEFAULT_SURFACES)

    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>FX Options Analytics</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>{CSS}</style></head><body>
<h1>FX Options Analytics</h1>
<div class="tab-bar">
<button class="tab-btn active" onclick="showTab('mktdata')">Market Data</button>
<button class="tab-btn" onclick="showTab('surface')">Surface Analysis</button>
<button class="tab-btn" onclick="showTab('portfolio')">Portfolio Analysis</button>
<button class="tab-btn" onclick="showTab('dtcc')">DTCC Live Surface</button>
</div>
'''
    # Tab 0: Market Data
    html += '''<div id="tab-mktdata" class="tab-content active">
<div style="display:flex;gap:12px;align-items:center;margin-bottom:12px;flex-wrap:wrap">
<div style="display:flex;gap:4px;flex-wrap:wrap" id="mkt-pair-tabs"></div>
<div class="add-pair-btn" onclick="addNewPair()">+ Add Pair</div>
<span style="margin-left:auto;font-size:11px;color:#888" id="mkt-status"></span>
</div>
<div id="mkt-pair-content"><div style="text-align:center;padding:40px;color:#888">Select or add a pair above to begin marking</div></div>
</div>
'''
    # Tab 1: Surface Analysis
    html += '''<div id="tab-surface" class="tab-content">
<div class="market-bar" id="surface-bar"></div>
<div class="grid"><div>
<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:10px">
<h2 style="margin:0" id="hm-title">Richness Score</h2>
<div><button class="btn-toggle active" onclick="setHmView('richness')" id="hm-btn-richness">Richness</button>
<button class="btn-toggle" onclick="setHmView('theta')" id="hm-btn-theta">Theta</button>
<button class="btn-toggle" onclick="setHmView('vega')" id="hm-btn-vega">Vega</button></div></div>
<div id="heatmap" style="height:400px"></div>
<div class="help">Richness = |Theta + Delta*T/N Roll| / Gamma, normalized 1-5</div></div>
<div class="card"><h2>Richness by Tenor</h2><h3>Cheapest Gamma</h3><div id="cheap-tbl"></div><h3>Richest Gamma</h3><div id="rich-tbl"></div></div>
</div><div>
<div class="card"><h2>Forward Points</h2><div id="fwd-tbl"></div></div>
<div class="card" style="margin-top:15px"><h2>Greeks Calculator</h2>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
<div><label style="font-size:12px">Tenor</label><select id="calc-t" style="width:100%"></select></div>
<div><label style="font-size:12px">Delta</label><select id="calc-d" style="width:100%"></select></div>
</div><button class="btn btn-primary" onclick="calcGreeks()" style="margin-top:10px">Calculate</button>
<div id="calc-result"></div></div>
</div></div></div>
'''
    # Tab 2: Portfolio
    html += '''<div id="tab-portfolio" class="tab-content">
<div id="port-pair-bar" style="margin-bottom:12px"></div>
<div class="summary-grid">
<div class="summary-box"><div class="summary-label">Total Gamma ($M)</div><div class="summary-value" id="sg">-</div></div>
<div class="summary-box"><div class="summary-label">Total Theta ($K/day)</div><div class="summary-value" id="st">-</div></div>
<div class="summary-box"><div class="summary-label">Wtd Richness</div><div class="summary-value" id="sr">-</div></div>
<div class="summary-box"><div class="summary-label">Total Vega ($K)</div><div class="summary-value" id="sv">-</div></div>
<div class="summary-box"><div class="summary-label">Projected Decay</div><div class="summary-value" id="sd">-</div></div>
</div>
<div class="portfolio-grid">
<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;gap:10px">
<h2 style="margin:0" id="ph-title">Portfolio Richness (1-5)</h2>
<div>
<button class="btn-toggle active" onclick="setPhView('richness')" id="ph-btn-richness">Richness</button>
<button class="btn-toggle" onclick="setPhView('vega')" id="ph-btn-vega">Vega</button>
<button class="btn-toggle" onclick="setPhView('decay')" id="ph-btn-decay">Decay</button>
</div></div>
<div id="port-hm" style="height:350px"></div>
<div class="help">Click a cell to see positions. Gamma-weighted avg richness.</div></div>
<div class="card"><h2>Position Details</h2><div id="pos-tbl" style="max-height:500px;overflow-y:auto"></div></div>
</div>
<div class="card" style="margin-top:15px"><h2>Inefficient Positions</h2>
<div style="margin-bottom:10px">
<button class="btn-toggle active" onclick="setIneff('long')" id="ie-long">Long (Rich Gamma)</button>
<button class="btn-toggle" onclick="setIneff('short')" id="ie-short">Short (Cheap Gamma)</button>
</div><div id="ineff-tbl" style="max-height:300px;overflow-y:auto"></div></div>
<div class="card"><h2>Portfolio Greeks Over Time</h2>
<div style="margin-bottom:10px">
<button class="btn-toggle active" onclick="setTv('gamma')" id="tv-gamma">Gamma</button>
<button class="btn-toggle" onclick="setTv('theta')" id="tv-theta">Theta</button>
<button class="btn-toggle" onclick="setTv('vega')" id="tv-vega">Vega</button>
<button class="btn-toggle" onclick="setTv('cumDecay')" id="tv-cumDecay">Cum Decay</button>
</div><div id="time-chart" style="height:350px"></div></div>
</div></div>
'''
    # Tab 3: DTCC
    html += '''<div id="tab-dtcc" class="tab-content">
<div class="dtcc-container" id="dtcc-root"><div style="text-align:center;padding:40px;color:#888">Click this tab to connect...</div></div>
</div>
'''

    # Script block — data injection + JS engine
    html += '<script>\n'
    html += f'var excelPositions={pos_json};\n'
    html += f'var defaultSurfaces={defaults_json};\n'
    html += f'var pairsInPositions={pairs_in_pos};\n'
    html += JS_ENGINE
    html += '\ninitApp();\n</script></body></html>'

    with open(output, 'w', encoding='utf-8') as fout:
        fout.write(html)
    return output

JS_ENGINE = r'''
// ============================================================
// GLOBAL STATE
// ============================================================
var mktSurfaces={};  // pair->{spot,r_d,r_f,tenors:[{tenor,T,atm,rr25,rr10,fly25,fly10,fwdPts}]}
var activeMktPair=null, activeAnPair=null, activePortPair=null;
var currentHm='richness', currentTv='gamma', tsData=null;
var TENORS=['O/N','1W','2W','1M','2M','3M','6M','9M','1Y','2Y'];
var TENOR_T=[1/365,7/365,14/365,1/12,2/12,3/12,6/12,9/12,1.0,2.0];
var DELTAS=['10P','15P','20P','25P','30P','35P','40P','45P','ATM','45C','40C','35C','30C','25C','20C','15C','10C'];
var DL=['10\u0394P','15\u0394P','20\u0394P','25\u0394P','30\u0394P','35\u0394P','40\u0394P','45\u0394P','ATM','45\u0394C','40\u0394C','35\u0394C','30\u0394C','25\u0394C','20\u0394C','15\u0394C','10\u0394C'];
var DV=[[-0.10,false],[-0.15,false],[-0.20,false],[-0.25,false],[-0.30,false],[-0.35,false],[-0.40,false],[-0.45,false],[0,null],[0.45,true],[0.40,true],[0.35,true],[0.30,true],[0.25,true],[0.20,true],[0.15,true],[0.10,true]];

// ============================================================
// BLACK-76 ENGINE
// ============================================================
function ncdf(x){var a1=0.254829592,a2=-0.284496736,a3=1.421413741,a4=-1.453152027,a5=1.061405429,p=0.3275911;var s=x<0?-1:1;x=Math.abs(x)/Math.sqrt(2);var t=1/(1+p*x);return 0.5*(1+s*(1-((((a5*t+a4)*t+a3)*t+a2)*t+a1)*t*Math.exp(-x*x)));}
function npdf(x){return Math.exp(-0.5*x*x)/Math.sqrt(2*Math.PI);}
function b76d1(F,K,T,s){if(T<=0||s<=0)return 0;return(Math.log(F/K)+0.5*s*s*T)/(s*Math.sqrt(T));}
function b76P(F,K,T,s,r,ic){if(T<=0)return ic?Math.max(F-K,0):Math.max(K-F,0);var d1=b76d1(F,K,T,s),d2=d1-s*Math.sqrt(T),df=Math.exp(-r*T);return ic?df*(F*ncdf(d1)-K*ncdf(d2)):df*(K*ncdf(-d2)-F*ncdf(-d1));}
function b76D(F,K,T,s,r,S,ic){if(T<=0)return 0;var d1=b76d1(F,K,T,s),df=Math.exp(-r*T),dff=df*F/S;return ic?dff*ncdf(d1):dff*(ncdf(d1)-1);}
function b76G(F,K,T,s,r,S){if(T<=0||s<=0)return 0;var d1=b76d1(F,K,T,s),df=Math.exp(-r*T),dff=df*F/S;return dff*npdf(d1)/(S*s*Math.sqrt(T));}
function b76T(F,K,T,s,r,ic){if(T<=0)return 0;var d1=b76d1(F,K,T,s),df=Math.exp(-r*T),p=b76P(F,K,T,s,r,ic);return(r*p-df*F*npdf(d1)*s/(2*Math.sqrt(T)))/365;}
function b76V(F,K,T,s,r){if(T<=0)return 0;var d1=b76d1(F,K,T,s),df=Math.exp(-r*T);return F*df*npdf(d1)*Math.sqrt(T)*0.01;}
function kFromDC(F,T,s,td,dff){var lo=F*0.3,hi=F*3;for(var i=0;i<50;i++){var m=(lo+hi)/2,d1=(Math.log(F/m)+0.5*s*s*T)/(s*Math.sqrt(T));if(dff*ncdf(d1)>td)lo=m;else hi=m;}return(lo+hi)/2;}
function kFromDP(F,T,s,td,dff){var lo=F*0.3,hi=F*3;for(var i=0;i<50;i++){var m=(lo+hi)/2,d1=(Math.log(F/m)+0.5*s*s*T)/(s*Math.sqrt(T));if(dff*(ncdf(d1)-1)<td)hi=m;else lo=m;}return(lo+hi)/2;}
function atmK(F,T,s){return F*Math.exp(0.5*s*s*T);}
function solveSmile(a,r25,f25,r10,f10){var v=a/100;return{ATM:v,'25C':v+f25/100+0.5*r25/100,'25P':v+f25/100-0.5*r25/100,'10C':v+f10/100+0.5*r10/100,'10P':v+f10/100-0.5*r10/100};}
function ivol(sm,d,ic){if(d===0)return sm.ATM;var a=Math.abs(d);if(ic){if(a<=0.1)return sm['10C'];if(a<=0.25)return sm['10C']+(a-0.1)/0.15*(sm['25C']-sm['10C']);return sm['25C']+(a-0.25)/0.25*(sm.ATM-sm['25C']);}else{if(a<=0.1)return sm['10P'];if(a<=0.25)return sm['10P']+(a-0.1)/0.15*(sm['25P']-sm['10P']);return sm['25P']+(a-0.25)/0.25*(sm.ATM-sm['25P']);}}

// Build full surface from market data marks
function buildSurf(pd){
    var S=pd.spot,rd=pd.r_d,N=1e6,rows=[],fwd=[];
    var onT=pd.tenors.find(function(t){return t.tenor==='O/N';});
    var tnR=onT?-onT.fwdPts:0;
    pd.tenors.forEach(function(tr){
        var T=tr.T,F=S+tr.fwdPts/10000,dfd=Math.exp(-rd*T),dff=dfd*F/S;
        var sm=solveSmile(tr.atm,tr.rr25,tr.fly25,tr.rr10,tr.fly10);
        var days=Math.max(1,Math.round(T*365));
        fwd.push({tenor:tr.tenor,T:T,days:days,forward:F,fwdPts:tr.fwdPts});
        DV.forEach(function(dv,di){
            var dval=dv[0],ic=dv[1],sig,K,th,delta;
            if(dval===0){sig=sm.ATM;K=atmK(F,T,sig);th=(b76T(F,K,T,sig,rd,true)+b76T(F,K,T,sig,rd,false))/2;delta=b76D(F,K,T,sig,rd,S,true);}
            else{sig=ivol(sm,dval,ic);K=ic?kFromDC(F,T,sig,Math.abs(dval),dff):kFromDP(F,T,sig,dval,dff);th=b76T(F,K,T,sig,rd,ic);delta=b76D(F,K,T,sig,rd,S,ic);}
            var gam=b76G(F,K,T,sig,rd,S),veg=b76V(F,K,T,sig,rd);
            var roll=delta*(tnR/10000),cost=gam>1e-12?Math.abs(th+roll)/gam:0;
            rows.push({tenor:tr.tenor,T:T,days:days,dl:DELTAS[di],dval:dval,strike:K,vol:sig*100,delta:delta,
                gamma:gam*S*0.01*N/1e6,theta:th*N/1000,decay:th*N*days/1000,vega:veg*N/1000,gcr:cost});
        });
    });
    var rc=rows.map(function(r){return r.gcr;}).filter(function(v){return v>0;}).sort(function(a,b){return a-b;});
    var p5=rc[Math.floor(rc.length*0.05)]||0,p95=rc[Math.floor(rc.length*0.95)]||1;
    rows.forEach(function(r){r.rich=r.gcr>0?Math.max(1,Math.min(5,1+(r.gcr-p5)/(p95-p5)*4)):1;});
    return{data:rows,fwd:fwd,p5:p5,p95:p95,tnR:tnR,spot:S,rd:rd,rf:pd.r_f};
}

function richColor(s){if(s<=3){var t=(s-1)/2;return'rgb('+Math.round(21+(224-21)*t)+','+Math.round(101+(224-101)*t)+','+Math.round(192+(224-192)*t)+')';}else{var t=(s-3)/2;return'rgb('+Math.round(224+(198-224)*t)+','+Math.round(224+(40-224)*t)+','+Math.round(224+(40-224)*t)+')';}}

// ============================================================
// TAB SWITCHING
// ============================================================
function showTab(tab){
    document.querySelectorAll('.tab-btn').forEach(function(b){b.classList.remove('active');});
    document.querySelectorAll('.tab-content').forEach(function(c){c.classList.remove('active');});
    document.querySelector('.tab-btn[onclick*="'+tab+'"]').classList.add('active');
    document.getElementById('tab-'+tab).classList.add('active');
    if(tab==='surface')renderSurfTab();
    if(tab==='portfolio')renderPortTab();
    if(tab==='dtcc')dtccInit();
    if(tab!=='dtcc'&&typeof dtccPause==='function')dtccPause();
}

// ============================================================
// INIT
// ============================================================
function initApp(){
    Object.keys(defaultSurfaces).forEach(function(p){mktSurfaces[p]=JSON.parse(JSON.stringify(defaultSurfaces[p]));});
    pairsInPositions.forEach(function(p){if(!mktSurfaces[p])mktSurfaces[p]={spot:1.0,r_d:0.04,r_f:0.02,tenors:TENORS.map(function(t,i){return{tenor:t,T:TENOR_T[i],atm:8,rr25:0,rr10:0,fly25:0.2,fly10:0.5,fwdPts:0};})};});
    if(Object.keys(mktSurfaces).length>0){activeMktPair=Object.keys(mktSurfaces)[0];activeAnPair=activeMktPair;activePortPair=pairsInPositions[0]||activeMktPair;}
    renderMktTabs();renderMktContent();
}

// ============================================================
// TAB 0: MARKET DATA
// ============================================================
function renderMktTabs(){
    var bar=document.getElementById('mkt-pair-tabs');var html='';
    Object.keys(mktSurfaces).forEach(function(p){html+='<div class="pair-tab'+(p===activeMktPair?' active':'')+'" onclick="selMktPair(\''+p+'\')">'+p+'</div>';});
    bar.innerHTML=html;
    document.getElementById('mkt-status').textContent=Object.keys(mktSurfaces).length+' pairs marked';
}
function selMktPair(p){activeMktPair=p;renderMktTabs();renderMktContent();}
function addNewPair(){var p=prompt('Enter pair (e.g. AUDUSD):');if(!p)return;p=p.toUpperCase().replace('/','').trim();if(p.length!==6){alert('Must be 6 chars');return;}if(mktSurfaces[p]){selMktPair(p);return;}mktSurfaces[p]={spot:1.0,r_d:0.04,r_f:0.02,tenors:TENORS.map(function(t,i){return{tenor:t,T:TENOR_T[i],atm:8,rr25:0,rr10:0,fly25:0.2,fly10:0.5,fwdPts:0};})};activeMktPair=p;renderMktTabs();renderMktContent();}

function renderMktContent(){
    if(!activeMktPair||!mktSurfaces[activeMktPair]){document.getElementById('mkt-pair-content').innerHTML='<div style="text-align:center;padding:40px;color:#888">Select a pair</div>';return;}
    var pd=mktSurfaces[activeMktPair];
    var html='<div class="card"><div class="mkt-param-row">';
    html+='<div class="mkt-param"><label>Spot</label><input class="mkt-input" id="mkt-spot" type="number" step="0.0001" value="'+pd.spot+'" onchange="onMktP()"></div>';
    html+='<div class="mkt-param"><label>Terms Rate (%)</label><input class="mkt-input" id="mkt-rd" type="number" step="0.001" value="'+(pd.r_d*100).toFixed(3)+'" onchange="onMktP()"></div>';
    html+='<div class="mkt-param"><label>Base Rate (%)</label><input class="mkt-input" id="mkt-rf" type="number" step="0.001" value="'+(pd.r_f*100).toFixed(3)+'" onchange="onMktP()"></div>';
    html+='</div></div>';
    html+='<div class="card"><h2 style="margin-bottom:10px">Vol Surface: '+activeMktPair+'</h2>';
    html+='<table class="vol-table"><thead><tr><th style="text-align:left">Tenor</th><th>ATM</th><th>RR25</th><th>RR10</th><th>FLY25</th><th>FLY10</th><th>Fwd Pts</th></tr></thead><tbody>';
    pd.tenors.forEach(function(t,i){
        html+='<tr><td>'+t.tenor+'</td>';
        ['atm','rr25','rr10','fly25','fly10','fwdPts'].forEach(function(f){
            html+='<td><input data-f="'+f+'" data-i="'+i+'" value="'+t[f].toFixed(2)+'" onchange="onVolC(this)"></td>';
        });
        html+='</tr>';
    });
    html+='</tbody></table><div class="help">Edit any cell — surface recomputes on tab switch. ATM/RR/FLY in vol %, Fwd Pts in pips.</div></div>';
    document.getElementById('mkt-pair-content').innerHTML=html;
}
function onVolC(el){var f=el.getAttribute('data-f'),i=parseInt(el.getAttribute('data-i')),v=parseFloat(el.value);if(!isNaN(v))mktSurfaces[activeMktPair].tenors[i][f]=v;}
function onMktP(){var pd=mktSurfaces[activeMktPair];pd.spot=parseFloat(document.getElementById('mkt-spot').value)||pd.spot;pd.r_d=(parseFloat(document.getElementById('mkt-rd').value)||0)/100;pd.r_f=(parseFloat(document.getElementById('mkt-rf').value)||0)/100;}

// ============================================================
// TAB 1: SURFACE ANALYSIS
// ============================================================
function renderSurfTab(){
    if(!activeAnPair)activeAnPair=Object.keys(mktSurfaces)[0];
    var pd=mktSurfaces[activeAnPair];if(!pd)return;
    var cs=buildSurf(pd);
    // Market bar
    var barH='<div class="market-item"><div class="label">Pair</div><select onchange="activeAnPair=this.value;renderSurfTab()" style="font-size:14px;font-weight:bold;color:#90caf9;background:#3d3d3d;border:1px solid #555;border-radius:4px;padding:4px 8px">';
    Object.keys(mktSurfaces).forEach(function(p){barH+='<option'+(p===activeAnPair?' selected':'')+'>'+p+'</option>';});
    barH+='</select></div>';
    var onF=pd.tenors.find(function(t){return t.tenor==='O/N';});
    barH+='<div class="market-item"><div class="label">Spot</div><div class="value">'+cs.spot.toFixed(4)+'</div></div>';
    barH+='<div class="market-item"><div class="label">Terms Rate</div><div class="value">'+(cs.rd*100).toFixed(2)+'%</div></div>';
    barH+='<div class="market-item"><div class="label">O/N Fwd Pts</div><div class="value">'+(onF?onF.fwdPts.toFixed(2):'—')+'</div></div>';
    barH+='<div class="market-item"><div class="label">Notional ($M)</div><input type="number" id="notional" class="notional-input" value="1" min="0.1" step="0.1" onchange="renderSurfTab()"></div>';
    document.getElementById('surface-bar').innerHTML=barH;
    // Heatmap
    renderHm(cs);renderTbls(cs);renderFwd(cs);popCalc();
}
function getN(){return parseFloat((document.getElementById('notional')||{}).value)||1;}
function setHmView(v){currentHm=v;['richness','theta','vega'].forEach(function(k){document.getElementById('hm-btn-'+k).className='btn-toggle'+(k===v?' active':'');});document.getElementById('hm-title').textContent={richness:'Richness Score',theta:'Theta ($K/day)',vega:'Vega ($K)'}[v];renderSurfTab();}
function renderHm(cs){
    var n=getN(),mat=[],z1=1,z5=5,bt='Score',clr=[[0,'#1565c0'],[0.5,'#e0e0e0'],[1,'#c62828']];
    TENORS.forEach(function(t){var row=DELTAS.map(function(d){var pt=cs.data.find(function(r){return r.tenor===t&&r.dl===d;});if(!pt)return null;if(currentHm==='richness')return pt.rich;if(currentHm==='theta')return Math.abs(pt.theta*n);return Math.abs(pt.vega*n);});mat.push(row);});
    if(currentHm!=='richness'){var fl=mat.flat().filter(function(v){return v!==null;});z1=fl.length?Math.min.apply(null,fl):0;z5=fl.length?Math.max.apply(null,fl):1;bt=currentHm==='theta'?'$K/day':'$K';}
    var txt=mat.map(function(r){return r.map(function(v){return v!==null?v.toFixed(2):'';});});
    Plotly.react('heatmap',[{z:mat,x:DL,y:TENORS,type:'heatmap',colorscale:clr,text:txt,texttemplate:'%{text}',textfont:{size:10,color:'black'},colorbar:{title:{text:bt,font:{color:'#e0e0e0'}},tickfont:{color:'#e0e0e0'},len:.9},zmin:z1,zmax:z5}],{margin:{t:20,b:80,l:60,r:50},xaxis:{title:'Delta',tickangle:45,color:'#e0e0e0'},yaxis:{title:'Tenor',color:'#e0e0e0'},paper_bgcolor:'#3d3d3d',plot_bgcolor:'#3d3d3d'},{displayModeBar:false,responsive:true});
}
function renderTbls(cs){
    var n=getN(),ch='<table><tr><th style="text-align:left">Tenor</th><th>Delta</th><th>Rich</th><th>\u03b8 ($K)</th><th>\u03b3 ($M)</th></tr>',rh=ch;
    TENORS.forEach(function(t){var pts=cs.data.filter(function(r){return r.tenor===t;});if(!pts.length)return;
        var c=pts.reduce(function(a,b){return a.rich<b.rich?a:b;}),r=pts.reduce(function(a,b){return a.rich>b.rich?a:b;});
        ch+='<tr class="cheap"><td>'+t+'</td><td>'+c.dl+'</td><td>'+c.rich.toFixed(2)+'</td><td>'+Math.abs(c.theta*n).toFixed(2)+'</td><td>'+Math.abs(c.gamma*n).toFixed(4)+'</td></tr>';
        rh+='<tr class="rich"><td>'+t+'</td><td>'+r.dl+'</td><td>'+r.rich.toFixed(2)+'</td><td>'+Math.abs(r.theta*n).toFixed(2)+'</td><td>'+Math.abs(r.gamma*n).toFixed(4)+'</td></tr>';
    });ch+='</table>';rh+='</table>';document.getElementById('cheap-tbl').innerHTML=ch;document.getElementById('rich-tbl').innerHTML=rh;
}
function renderFwd(cs){var h='<table><tr><th style="text-align:left">Tenor</th><th>Days</th><th>Forward</th><th>Fwd Pts</th></tr>';cs.fwd.forEach(function(f){h+='<tr><td>'+f.tenor+'</td><td>'+f.days+'</td><td>'+f.forward.toFixed(5)+'</td><td>'+f.fwdPts.toFixed(2)+'</td></tr>';});h+='</table>';document.getElementById('fwd-tbl').innerHTML=h;}
function popCalc(){var st=document.getElementById('calc-t');st.innerHTML='';TENORS.forEach(function(t){st.innerHTML+='<option>'+t+'</option>';});var sd=document.getElementById('calc-d');sd.innerHTML='';DELTAS.forEach(function(d,i){sd.innerHTML+='<option value="'+d+'">'+DL[i]+'</option>';});}
function calcGreeks(){var pd=mktSurfaces[activeAnPair];if(!pd)return;var cs=buildSurf(pd);var t=document.getElementById('calc-t').value,d=document.getElementById('calc-d').value;var pt=cs.data.find(function(r){return r.tenor===t&&r.dl===d;});if(!pt)return;var n=getN();document.getElementById('calc-result').innerHTML='<div style="background:#2d2d2d;padding:12px;border-radius:6px;margin-top:10px"><div style="margin-bottom:10px"><strong>'+pt.tenor+' '+pt.dl+'</strong> | K:'+pt.strike.toFixed(4)+' | Vol:'+pt.vol.toFixed(2)+'% | Days:'+pt.days+'</div><div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;text-align:center"><div style="padding:8px;background:#3d3d3d;border-radius:4px"><div style="font-size:10px;color:#aaa">Delta</div><div style="font-size:14px;font-weight:bold;color:#90caf9">'+Math.abs(pt.delta).toFixed(4)+'</div></div><div style="padding:8px;background:#3d3d3d;border-radius:4px"><div style="font-size:10px;color:#aaa">Gamma ($M)</div><div style="font-size:14px;font-weight:bold;color:#90caf9">'+Math.abs(pt.gamma*n).toFixed(4)+'</div></div><div style="padding:8px;background:#3d3d3d;border-radius:4px"><div style="font-size:10px;color:#aaa">Theta ($K)</div><div style="font-size:14px;font-weight:bold;color:#90caf9">'+Math.abs(pt.theta*n).toFixed(2)+'</div></div><div style="padding:8px;background:#3d3d3d;border-radius:4px"><div style="font-size:10px;color:#aaa">Vega ($K)</div><div style="font-size:14px;font-weight:bold;color:#90caf9">'+Math.abs(pt.vega*n).toFixed(2)+'</div></div></div></div>';}

// ============================================================
// TAB 2: PORTFOLIO
// ============================================================
function renderPortTab(){
    if(!activePortPair)activePortPair=pairsInPositions[0]||Object.keys(mktSurfaces)[0];
    var bh='<span style="font-size:12px;color:#aaa;margin-right:8px">Pair:</span>';
    pairsInPositions.forEach(function(p){bh+='<button class="btn-toggle'+(p===activePortPair?' active':'')+'" onclick="activePortPair=\''+p+'\';renderPortTab()">'+p+'</button>';});
    document.getElementById('port-pair-bar').innerHTML=bh;
    var pd=mktSurfaces[activePortPair];
    if(!pd){document.getElementById('pos-tbl').innerHTML='<div style="text-align:center;padding:40px;color:#aaa">No surface for '+activePortPair+'. Mark it in Market Data tab.</div>';return;}
    var cs=buildSurf(pd);
    var pos=excelPositions.filter(function(p){return p.pair===activePortPair;});
    if(!pos.length){document.getElementById('pos-tbl').innerHTML='<div style="text-align:center;padding:40px;color:#aaa">No positions for '+activePortPair+'</div>';return;}
    var comp=[];
    pos.forEach(function(p,i){var ic=p.type==='C';var g=cpg(cs,p.strike,p.T,p.vol,p.notional,ic);comp.push({id:i+1,strike:p.strike,expiry:p.expiry,vol:p.vol,notional:p.notional,type:p.type,T:p.T,days:p.days,delta:g.delta,gamma:g.gamma,theta:g.theta,vega:g.vega,decay:g.decay,rich:g.rich});});
    var tG=0,tT=0,tV=0,rS=0,gA=0;
    comp.forEach(function(p){tG+=p.gamma;tT+=p.theta;tV+=p.vega;var ga=Math.abs(p.gamma);rS+=p.rich*ga;gA+=ga;});
    document.getElementById('sg').textContent=tG.toFixed(4);
    document.getElementById('st').textContent=(tT<0?'Pay $':'Rcv $')+Math.abs(tT).toFixed(2)+'K';document.getElementById('st').style.color=tT<0?'#ef5350':'#66bb6a';
    document.getElementById('sr').textContent=gA>0?(rS/gA).toFixed(2):'-';
    document.getElementById('sv').textContent=tV.toFixed(2);
    var h='<table><tr><th>#</th><th>Strike</th><th>Expiry</th><th>Days</th><th>Vol</th><th>Notl</th><th>Type</th><th>\u03b3</th><th>\u03b8</th><th>\u03bd</th><th>Rich</th></tr>';
    comp.forEach(function(p){h+='<tr class="'+(p.notional>=0?'cheap':'rich')+'"><td>'+p.id+'</td><td>'+p.strike.toFixed(3)+'</td><td>'+p.expiry+'</td><td>'+p.days+'</td><td>'+p.vol.toFixed(1)+'</td><td>'+(p.notional>=0?'+':'')+p.notional.toFixed(1)+'</td><td>'+p.type+'</td><td>'+p.gamma.toFixed(4)+'</td><td>'+p.theta.toFixed(2)+'</td><td>'+p.vega.toFixed(1)+'</td><td><span class="richness-badge" style="background:'+richColor(p.rich)+'">'+p.rich.toFixed(2)+'</span></td></tr>';});
    h+='</table>';document.getElementById('pos-tbl').innerHTML=h;
    portComp=comp;portCS=cs;
    buildPortHm(cs,comp);renderIneff();buildTS(cs,comp);
}
var portComp=null,portCS=null,currentPh='richness',currentIneff='long';

function setPhView(v){currentPh=v;['richness','vega','decay'].forEach(function(k){document.getElementById('ph-btn-'+k).className='btn-toggle'+(k===v?' active':'');});
    document.getElementById('ph-title').textContent={richness:'Portfolio Richness (1-5)',vega:'Portfolio Vega ($K)',decay:'Portfolio Decay ($K/day)'}[v];renderPortHm();}

function buildPortHm(cs,comp){
    var tenorList=TENORS;
    var allK=comp.map(function(p){return p.strike;});var mnK=Math.min.apply(null,allK),mxK=Math.max.apply(null,allK),rng=mxK-mnK;
    var maxB=18,rw=rng/maxB;var nice=[0.001,0.0025,0.005,0.01,0.025,0.05,0.1,0.25,0.5,1,5,10,50,100,500];
    var bw=nice.find(function(w){return w>=rw;})||500;
    var uK=[].concat(allK).filter(function(v,i,a){return a.indexOf(v)===i;}).sort(function(a,b){return a-b;});
    var useB=uK.length>maxB;var sB,sL;
    if(useB){var s=Math.floor(mnK/bw)*bw,e=Math.ceil(mxK/bw)*bw;sB=[];sL=[];for(var b=s;b<e+bw;b+=bw){sB.push(b);sL.push(b.toFixed(3));}}
    else{sB=uK;sL=uK.map(function(k){return k.toFixed(3);});}
    var nb=sB.length;
    portHmData={strikes:sB,labels:sL,bw:useB?bw:0,
        rich:tenorList.map(function(){return sB.map(function(){return null;});}),
        vega:tenorList.map(function(){return sB.map(function(){return 0;});}),
        decay:tenorList.map(function(){return sB.map(function(){return 0;});}),
        dir:tenorList.map(function(){return sB.map(function(){return null;});}),
        byBucket:tenorList.map(function(){return sB.map(function(){return[];});})};
    function getSI(k){if(!useB)return sB.indexOf(k);return Math.max(0,Math.min(Math.floor((k-sB[0])/bw),nb-1));}
    function getTI(days){var bk=[['O/N',1],['1W',7],['2W',14],['1M',30],['2M',61],['3M',91],['6M',182],['9M',274],['1Y',365],['2Y',730]];for(var i=bk.length-1;i>=0;i--){if(days>=bk[i][1]*0.7)return tenorList.indexOf(bk[i][0]);}return 0;}
    var gSum=tenorList.map(function(){return sB.map(function(){return 0;});}),rgSum=tenorList.map(function(){return sB.map(function(){return 0;});}),nSum=tenorList.map(function(){return sB.map(function(){return 0;});});
    comp.forEach(function(p){var ti=getTI(p.days),si=getSI(p.strike);if(ti<0||si<0)return;
        portHmData.vega[ti][si]+=p.vega;portHmData.decay[ti][si]+=p.theta;portHmData.byBucket[ti][si].push(p);
        var ga=Math.abs(p.gamma);gSum[ti][si]+=ga;rgSum[ti][si]+=p.rich*ga;nSum[ti][si]+=p.notional;});
    for(var ti=0;ti<tenorList.length;ti++){for(var si=0;si<nb;si++){if(gSum[ti][si]>1e-10){portHmData.rich[ti][si]=rgSum[ti][si]/gSum[ti][si];portHmData.dir[ti][si]=nSum[ti][si]>=0?'L':'S';}}}
    renderPortHm();
}
var portHmData=null;
function renderPortHm(){
    if(!portHmData)return;var M,cs2,z1,z5,bt;
    if(currentPh==='richness'){M=portHmData.rich;cs2=[[0,'#1565c0'],[0.5,'#e0e0e0'],[1,'#c62828']];z1=1;z5=5;bt='Score';}
    else if(currentPh==='vega'){M=portHmData.vega.map(function(r){return r.map(function(v){return v===0?null:v;});});cs2=[[0,'#1565c0'],[0.5,'#e0e0e0'],[1,'#c62828']];bt='$K';var fl=M.flat().filter(function(v){return v!==null;});var mx=fl.length>0?Math.max(Math.abs(Math.min.apply(null,fl)),Math.abs(Math.max.apply(null,fl))):1;z1=-mx;z5=mx;}
    else{M=portHmData.decay.map(function(r){return r.map(function(v){return v===0?null:v;});});cs2=[[0,'#1565c0'],[0.5,'#e0e0e0'],[1,'#c62828']];bt='$K/day';var fl=M.flat().filter(function(v){return v!==null;});var mx=fl.length>0?Math.max(Math.abs(Math.min.apply(null,fl)),Math.abs(Math.max.apply(null,fl))):1;z1=-mx;z5=mx;}
    var dir=portHmData.dir;
    var txt=M.map(function(r,ti){return r.map(function(v,si){if(v===null)return'';var d=dir[ti]&&dir[ti][si]?dir[ti][si]:'';return d+' '+v.toFixed(currentPh==='vega'?1:2);});});
    Plotly.react('port-hm',[{z:M,x:portHmData.labels,y:TENORS,type:'heatmap',colorscale:cs2,text:txt,texttemplate:'%{text}',textfont:{size:9,color:'black'},colorbar:{title:{text:bt,font:{color:'#e0e0e0'}},tickfont:{color:'#e0e0e0'},len:.9},zmin:z1,zmax:z5,hoverongaps:false}],
    {margin:{t:20,b:80,l:60,r:50},xaxis:{title:'Strike',tickangle:45,color:'#e0e0e0',type:'category'},yaxis:{title:'Tenor',color:'#e0e0e0'},paper_bgcolor:'#3d3d3d',plot_bgcolor:'#3d3d3d'},{displayModeBar:false,responsive:true});
    var hm=document.getElementById('port-hm');if(hm.removeAllListeners)hm.removeAllListeners('plotly_click');
    hm.on('plotly_click',function(data){var pt=data.points[0],ti=TENORS.indexOf(pt.y),si=pt.pointIndex[1];if(ti>=0&&si>=0)showDrill(ti,si);});
}

function showDrill(ti,si){
    if(!portHmData)return;var pos=portHmData.byBucket[ti][si];
    var bw=portHmData.bw,sk=portHmData.strikes[si];
    var skL=bw>0?sk.toFixed(3)+'-'+(sk+bw).toFixed(3):sk.toFixed(3);
    document.getElementById('drill-title').textContent=TENORS[ti]+' / '+skL+' ('+pos.length+')';
    if(!pos.length){document.getElementById('drill-content').innerHTML='<p style="text-align:center;color:#aaa">No positions</p>';}
    else{var h='<table><tr><th>#</th><th>Strike</th><th>Expiry</th><th>Vol</th><th>Notl</th><th>Type</th><th>\u03b3</th><th>\u03b8</th><th>\u03bd</th><th>Rich</th></tr>';
    pos.forEach(function(p){h+='<tr class="'+(p.notional>=0?'cheap':'rich')+'"><td>'+p.id+'</td><td>'+p.strike.toFixed(3)+'</td><td>'+p.expiry+'</td><td>'+p.vol.toFixed(1)+'</td><td>'+(p.notional>=0?'+':'')+p.notional.toFixed(1)+'</td><td>'+p.type+'</td><td>'+p.gamma.toFixed(4)+'</td><td>'+p.theta.toFixed(2)+'</td><td>'+p.vega.toFixed(1)+'</td><td><span class="richness-badge" style="background:'+richColor(p.rich)+'">'+p.rich.toFixed(2)+'</span></td></tr>';});
    h+='</table>';document.getElementById('drill-content').innerHTML=h;}
    document.getElementById('drill-modal').classList.add('show');
}
function closeDrillDown(){document.getElementById('drill-modal').classList.remove('show');}

function setIneff(v){currentIneff=v;['long','short'].forEach(function(k){document.getElementById('ie-'+k).className='btn-toggle'+(k===v?' active':'');});renderIneff();}
function renderIneff(){
    if(!portComp)return;var filtered,sorted;
    if(currentIneff==='long'){filtered=portComp.filter(function(p){return p.notional>0;});sorted=filtered.sort(function(a,b){return b.rich-a.rich;});}
    else{filtered=portComp.filter(function(p){return p.notional<0;});sorted=filtered.sort(function(a,b){return b.rich-a.rich;});}
    if(!sorted.length){document.getElementById('ineff-tbl').innerHTML='<p style="text-align:center;color:#aaa">No '+(currentIneff==='long'?'long':'short')+' positions</p>';return;}
    var h='<table><tr><th>#</th><th>Strike</th><th>Expiry</th><th>Days</th><th>Notl</th><th>Type</th><th>\u03b3</th><th>\u03b8</th><th>Decay</th><th>Rich</th></tr>';
    sorted.forEach(function(p){h+='<tr class="'+(p.notional>=0?'cheap':'rich')+'"><td>'+p.id+'</td><td>'+p.strike.toFixed(3)+'</td><td>'+p.expiry+'</td><td>'+p.days+'</td><td>'+(p.notional>=0?'+':'')+p.notional.toFixed(1)+'</td><td>'+p.type+'</td><td>'+p.gamma.toFixed(4)+'</td><td>'+p.theta.toFixed(2)+'</td><td>'+p.decay.toFixed(1)+'</td><td><span class="richness-badge" style="background:'+richColor(p.rich)+'">'+p.rich.toFixed(2)+'</span></td></tr>';});
    h+='</table>';document.getElementById('ineff-tbl').innerHTML=h;
}
function cpg(cs,K,T,vol,notional,ic){
    var S=cs.spot,rd=cs.rd,sig=vol/100;if(T<=0)T=1/365;
    var F=S;cs.fwd.forEach(function(fd,i){if(i<cs.fwd.length-1&&T>=cs.fwd[i].T&&T<=cs.fwd[i+1].T){var w=(T-cs.fwd[i].T)/(cs.fwd[i+1].T-cs.fwd[i].T);F=cs.fwd[i].forward+(cs.fwd[i+1].forward-cs.fwd[i].forward)*w;}});
    if(T<=cs.fwd[0].T)F=cs.fwd[0].forward;if(T>=cs.fwd[cs.fwd.length-1].T)F=cs.fwd[cs.fwd.length-1].forward;
    var dfd=Math.exp(-rd*T),dff=dfd*F/S,d1=(Math.log(F/K)+0.5*sig*sig*T)/(sig*Math.sqrt(T));
    var dr=ic?dff*ncdf(d1):dff*(ncdf(d1)-1),gr=dff*npdf(d1)/(S*sig*Math.sqrt(T));
    var pr=b76P(F,K,T,sig,rd,ic),tp=(rd*pr-dfd*F*npdf(d1)*sig/(2*Math.sqrt(T)))/365;
    var vr=F*dfd*npdf(d1)*Math.sqrt(T),rp=dr*(cs.tnR/10000);
    var gcr=gr>1e-12?Math.abs(tp+rp)/gr:0,rich=gcr>0?Math.max(1,Math.min(5,1+(gcr-cs.p5)/(cs.p95-cs.p5)*4)):1;
    var sign=notional>=0?1:-1,aN=Math.abs(notional);
    return{delta:dr*sign,gamma:gr*S*0.01*aN*sign,theta:tp*aN*1000*sign,vega:vr*0.01*aN*1000*sign,decay:tp*aN*1000*Math.max(1,Math.round(T*365))*sign,rich:rich};
}
function setTv(v){currentTv=v;['gamma','theta','vega','cumDecay'].forEach(function(k){document.getElementById('tv-'+k).className='btn-toggle'+(k===v?' active':'');});renderTS();}
function buildTS(cs,comp){
    if(!comp.length){tsData=null;return;}var mx=Math.min(Math.max.apply(null,comp.map(function(p){return p.days;})),365);
    var dt=[],gs=[],ts=[],vs=[],cd=[];var cum=0;
    for(var d=0;d<=mx;d++){dt.push(new Date(Date.now()+d*864e5).toISOString().slice(0,10));var g=0,t=0,v=0;
        comp.forEach(function(p){var rT=p.T-d/365;if(rT<1/365)return;var r=cpg(cs,p.strike,rT,p.vol,p.notional,p.type==='C');g+=r.gamma;t+=r.theta;v+=r.vega;});
        gs.push(g);ts.push(t);vs.push(v);cum+=t;cd.push(cum);}
    tsData={dt:dt,g:gs,t:ts,v:vs,cd:cd};
    if(cd.length){var fd=cd[cd.length-1];document.getElementById('sd').textContent=(fd<0?'Pay ':'Rcv ')+'$'+Math.abs(fd).toFixed(1)+'K';document.getElementById('sd').style.color=fd<0?'#ef5350':'#66bb6a';}
    renderTS();
}
function renderTS(){if(!tsData)return;var d=tsData,vm={gamma:{y:d.g,c:'#64b5f6'},theta:{y:d.t,c:'#ef5350'},vega:{y:d.v,c:'#ab47bc'},cumDecay:{y:d.cd,c:'#ffa726'}};var v=vm[currentTv];Plotly.react('time-chart',[{x:d.dt,y:v.y,type:'scatter',mode:'lines',line:{color:v.c,width:2},fill:(currentTv==='cumDecay'||currentTv==='theta')?'tozeroy':'none'}],{margin:{t:20,b:50,l:60,r:30},xaxis:{color:'#e0e0e0',gridcolor:'#555',type:'date'},yaxis:{color:'#e0e0e0',gridcolor:'#555',zeroline:true,zerolinecolor:'#888'},paper_bgcolor:'#3d3d3d',plot_bgcolor:'#3d3d3d',showlegend:false},{displayModeBar:false,responsive:true});}
'''

JS_DTCC = r'''
// ============================================================
// TAB 3: DTCC LIVE SURFACE
// ============================================================
var DTCC_G10=['EURUSD','USDJPY','GBPUSD','USDCHF','AUDUSD','NZDUSD','USDCAD','USDSEK','USDNOK','EURGBP','EURJPY','GBPJPY'];
var DTCC_EM=['USDMXN','USDBRL','USDTRY','USDZAR','USDCNH','USDINR','USDKRW','USDSGD','USDTWD','USDIDR','USDPHP','USDCLP','USDCOP','USDPEN'];
var DT=['O/N','1W','2W','1M','2M','3M','6M','9M','1Y','2Y'],DTM=[2,9,18,45,75,120,225,315,450,1500],DDL=['10\u0394P','25\u0394P','ATM','25\u0394C','10\u0394C'];
var DPA={'EUR/USD':'EURUSD','USD/EUR':'EURUSD','USD/JPY':'USDJPY','JPY/USD':'USDJPY','GBP/USD':'GBPUSD','USD/GBP':'GBPUSD','USD/CHF':'USDCHF','AUD/USD':'AUDUSD','USD/AUD':'AUDUSD','NZD/USD':'NZDUSD','USD/NZD':'NZDUSD','USD/CAD':'USDCAD','USD/SEK':'USDSEK','USD/NOK':'USDNOK','EUR/GBP':'EURGBP','EUR/JPY':'EURJPY','GBP/JPY':'GBPJPY','USD/MXN':'USDMXN','USD/BRL':'USDBRL','USD/TRY':'USDTRY','USD/ZAR':'USDZAR','USD/CNH':'USDCNH','USD/CNY':'USDCNH','USD/INR':'USDINR','USD/KRW':'USDKRW','USD/SGD':'USDSGD','USD/TWD':'USDTWD','USD/IDR':'USDIDR','USD/PHP':'USDPHP','USD/CLP':'USDCLP','USD/COP':'USDCOP','USD/PEN':'USDPEN','CHF/USD':'USDCHF','CAD/USD':'USDCAD','SEK/USD':'USDSEK','NOK/USD':'USDNOK','GBP/EUR':'EURGBP','JPY/EUR':'EURJPY','JPY/GBP':'GBPJPY','MXN/USD':'USDMXN','BRL/USD':'USDBRL','TRY/USD':'USDTRY','ZAR/USD':'USDZAR','CNH/USD':'USDCNH','INR/USD':'USDINR','KRW/USD':'USDKRW','SGD/USD':'USDSGD','TWD/USD':'USDTWD','IDR/USD':'USDIDR','PHP/USD':'USDPHP','CLP/USD':'USDCLP','COP/USD':'USDCOP','PEN/USD':'USDPEN'};
var dAP='EURUSD',dAT={},dCS={},dSS=null,dST=null,dSV='vol',dTV='atm',dTimer=null,dInited=false,dPM={};

function dtccGetBase(pair,tn){var pd=mktSurfaces[pair];if(!pd)return null;var tr=pd.tenors.find(function(t){return t.tenor===tn;});if(!tr)return null;var a=tr.atm,r25=tr.rr25||0,r10=tr.rr10||0,f25=Math.max(0,tr.fly25||0),f10=Math.max(0,tr.fly10||0);return{atm:a,rr25:r25,rr10:r10,fly25:f25,fly10:f10,vols:[a+f10-r10/2,a+f25-r25/2,a,a+f25+r25/2,a+f10+r10/2]};}
function dtccGenBase(a){var f25=Math.max(0.1,a*0.025),f10=Math.max(0.25,a*0.07);return{atm:a,rr25:0,rr10:0,fly25:f25,fly10:f10,vols:[a+f10,a+f25,a,a+f25,a+f10]};}
function dtccShifts(b,acc){var sh={dA:0,dR25:0,dR10:0,dF25:0,dF10:0},h=[],o=[];for(var i=0;i<5;i++){h[i]=acc[i].count>0;o[i]=h[i]?acc[i].sumIV/acc[i].sumW:null;}
    if(h[2])sh.dA=o[2]-b.atm;else{var n=0,d=0;for(var i=0;i<5;i++){if(!h[i])continue;n+=(o[i]-b.vols[i])*acc[i].sumW;d+=acc[i].sumW;}if(d>0)sh.dA=n/d;}
    if(h[1]&&h[3])sh.dR25=(o[3]-o[1])-b.rr25;else if(h[1]||h[3]){var ref=h[2]?o[2]:(b.atm+sh.dA);if(h[3])sh.dR25=2*(o[3]-ref-b.fly25)-b.rr25;else sh.dR25=2*(ref+b.fly25-o[1])-b.rr25;}
    if(h[0]&&h[4])sh.dR10=(o[4]-o[0])-b.rr10;else if(h[0]||h[4]){var ref=h[2]?o[2]:(b.atm+sh.dA);if(h[4])sh.dR10=2*(o[4]-ref-b.fly10)-b.rr10;else sh.dR10=2*(ref+b.fly10-o[0])-b.rr10;}
    if(sh.dR25!==0&&sh.dR10===0&&b.rr25!==0)sh.dR10=sh.dR25*(b.rr10/b.rr25);else if(sh.dR10!==0&&sh.dR25===0&&b.rr10!==0)sh.dR25=sh.dR10*(b.rr25/b.rr10);
    if(h[1]&&h[3]&&h[2])sh.dF25=(o[3]+o[1])/2-o[2]-b.fly25;if(h[0]&&h[4]&&h[2])sh.dF10=(o[4]+o[0])/2-o[2]-b.fly10;
    if(sh.dF25!==0&&sh.dF10===0&&b.fly25>0)sh.dF10=sh.dF25*(b.fly10/b.fly25);else if(sh.dF10!==0&&sh.dF25===0&&b.fly10>0)sh.dF25=sh.dF10*(b.fly25/b.fly10);
    return sh;}
function dtccApply(b,sh){var nA=b.atm+sh.dA,nR25=b.rr25+sh.dR25,nR10=b.rr10+sh.dR10,nF25=Math.max(0,b.fly25+sh.dF25),nF10=Math.max(0,b.fly10+sh.dF10);if(nF10<nF25)nF10=nF25;
    var r=[nA+nF10-nR10/2,nA+nF25-nR25/2,nA,nA+nF25+nR25/2,nA+nF10+nR10/2];for(var i=0;i<5;i++){if(i!==2&&r[i]<r[2])r[i]=r[2]+0.01;}return r;}
function dDB(d){var a=Math.abs(d);if(a<0.07)return-1;if(a<0.175)return d<0?0:4;if(a<0.375)return d<0?1:3;return 2;}
function dTB(d){for(var i=0;i<DTM.length;i++){if(d<=DTM[i])return i;}return DT.length-1;}
function dCD(F,K,s,T,ic){if(T<=0||s<=0||F<=0||K<=0)return ic?0.5:-0.5;var d1=(Math.log(F/K)+0.5*s*s*T)/(s*Math.sqrt(T));return ic?ncdf(d1):ncdf(d1)-1;}
function dGC(){var cc=new Set();DTCC_G10.concat(DTCC_EM).forEach(function(p){cc.add(p.slice(0,3));cc.add(p.slice(3));});cc.delete('USD');return Array.from(cc).join(',');}

function dtccInit(){
    if(!dInited){dInited=true;var root=document.getElementById('dtcc-root');root.innerHTML=dtccHTML();dtccBuildPB();}
    dtccFetch();if(dTimer)clearInterval(dTimer);dTimer=setInterval(function(){var el=document.getElementById('dtcc-auto');if(el&&el.checked)dtccFetch();},30000);
}
function dtccPause(){if(dTimer){clearInterval(dTimer);dTimer=null;}}

function dtccHTML(){return '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px"><div style="display:flex;align-items:center;gap:4px;font-size:12px;color:#aaa"><span class="dtcc-dot loading" id="dd"></span><span id="dt">Connecting...</span></div><div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap"><span style="font-size:11px;color:#888">Lookback</span><select style="padding:4px 8px;background:#3d3d3d;color:#e0e0e0;border:1px solid #555;border-radius:4px;font-size:12px" id="dlb" onchange="dtccFetch()"><option value="60">1h</option><option value="120">2h</option><option value="240" selected>4h</option><option value="480">8h</option><option value="1440">Day</option></select><span style="font-size:11px;color:#888">Min</span><select style="padding:4px 8px;background:#3d3d3d;color:#e0e0e0;border:1px solid #555;border-radius:4px;font-size:12px" id="dms" onchange="dtccFetch()"><option value="0">All</option><option value="5000000">$5M+</option><option value="10000000" selected>$10M+</option><option value="20000000">$20M+</option><option value="50000000">$50M+</option></select><button class="btn-toggle" onclick="dtccSnap()">Snap</button><button class="btn-toggle" onclick="dtccClrSnap()" style="color:#ef5350">Clear</button><label style="font-size:11px;color:#888;display:flex;align-items:center;gap:3px"><input type="checkbox" id="dtcc-auto" checked style="accent-color:#90caf9">Auto</label></div></div><div class="dtcc-pair-bar" id="dpb"></div><div class="dtcc-summary-row"><div class="dtcc-summary-badge"><div class="dl">Trades</div><div class="dv" id="ds-t">&mdash;</div><div class="ds" id="ds-p">&mdash;</div></div><div class="dtcc-summary-badge"><div class="dl">Points</div><div class="dv" id="ds-pt">&mdash;</div></div><div class="dtcc-summary-badge"><div class="dl">ATM Front</div><div class="dv" id="ds-af">&mdash;</div><div class="ds" id="ds-afc">&mdash;</div></div><div class="dtcc-summary-badge"><div class="dl">ATM 1Y</div><div class="dv" id="ds-a1">&mdash;</div><div class="ds" id="ds-a1c">&mdash;</div></div></div><div class="dtcc-grid"><div><div class="card"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px"><h2 style="margin:0;font-size:16px">Vol Surface</h2><div><button class="btn-toggle active" id="dsv-vol" onclick="dSetSV(\'vol\')">Vol</button><button class="btn-toggle" id="dsv-chg" onclick="dSetSV(\'chg\')">Chg</button><span class="dtcc-snap-badge" id="dsn">No snap</span></div></div><div id="dsp" style="height:320px"></div></div><div class="card" style="margin-top:12px"><h2 style="margin:0;font-size:16px;margin-bottom:8px">ATM / RR / Fly</h2><div style="overflow-x:auto"><table class="dtcc-strat-table"><thead><tr><th style="text-align:left">Tenor</th><th>ATM</th><th>&Delta;</th><th>25dRR</th><th>&Delta;</th><th>10dRR</th><th>&Delta;</th><th>25dFly</th><th>&Delta;</th><th>10dFly</th><th>&Delta;</th><th>#</th></tr></thead><tbody id="dsb"></tbody></table></div></div><div class="card" style="margin-top:12px"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px"><h2 style="margin:0;font-size:16px">Term Structure</h2><div><button class="btn-toggle active" id="dtv-atm" onclick="dSetTV(\'atm\')">ATM</button><button class="btn-toggle" id="dtv-rr25" onclick="dSetTV(\'rr25\')">25dRR</button><button class="btn-toggle" id="dtv-fly25" onclick="dSetTV(\'fly25\')">25dFly</button></div></div><div id="dtp" style="height:200px"></div></div></div><div><div class="card" style="height:100%;display:flex;flex-direction:column"><div style="display:flex;justify-content:space-between;align-items:center;padding-bottom:8px;border-bottom:1px solid #555"><h2 style="margin:0;font-size:16px">Prints: <span id="dfp">EUR/USD</span></h2><span style="font-size:11px;color:#888" id="dfc">0</span></div><div class="dtcc-feed-header"><span>Time</span><span>Type</span><span>Tenor</span><span>Vol</span><span>Notl</span><span>&Delta;</span></div><div class="dtcc-feed-panel" id="dfl" style="flex:1"></div></div></div></div>';}

function dtccBuildPB(){var bar=document.getElementById('dpb');if(!bar)return;var h='<span class="dtcc-pair-label">G10</span>';
    DTCC_G10.forEach(function(p){var hm=mktSurfaces[p]?'border-bottom:2px solid #66bb6a':'';h+='<div class="dtcc-pair-chip'+(p===dAP?' active':'')+'" onclick="dSP(\''+p+'\')" style="'+hm+'">'+p.slice(0,3)+'/'+p.slice(3)+'</div>';});
    h+='<div class="dtcc-pair-divider"></div><span class="dtcc-pair-label">EM</span>';
    DTCC_EM.forEach(function(p){var hm=mktSurfaces[p]?'border-bottom:2px solid #66bb6a':'';h+='<div class="dtcc-pair-chip'+(p===dAP?' active':'')+'" onclick="dSP(\''+p+'\')" style="'+hm+'">'+p.slice(0,3)+'/'+p.slice(3)+'</div>';});
    bar.innerHTML=h;}
function dSP(p){dAP=p;dtccBuildPB();document.getElementById('dfp').textContent=p.slice(0,3)+'/'+p.slice(3);dRA();}

async function dtccFetch(){var dot=document.getElementById('dd'),txt=document.getElementById('dt');if(!dot)return;dot.className='dtcc-dot loading';txt.textContent='Fetching...';
    var m=document.getElementById('dlb').value,ms=document.getElementById('dms').value,qs='ccys='+encodeURIComponent(dGC())+'&min_size='+ms+'&minutes='+m;
    var urls=['/api/optionflow?'+qs,'https://dtcc.ericlanalytics.com/api/optionflow?'+qs];var data=null;
    for(var i=0;i<urls.length;i++){try{var r=await fetch(urls[i]);if(!r.ok)throw new Error('HTTP '+r.status);data=await r.json();break;}catch(e){if(i===urls.length-1){dot.className='dtcc-dot err';txt.textContent='No connection';return;}}}
    if(!data||!data.trades){dot.className='dtcc-dot err';txt.textContent='Bad data';return;}
    dot.className='dtcc-dot ok';txt.textContent=(data.count||data.trades.length)+' trades';
    dProcTrades(data.trades);dBuildSurf();dRA();}

function dProcTrades(trades){dAT={};dPM={};DTCC_G10.concat(DTCC_EM).forEach(function(p){dAT[p]=[];});
    trades.forEach(function(t){try{var pk=DPA[t.pair];if(!pk||!dAT[pk])return;var iv=null;if(t.iv!=null&&t.iv!=='\u2014'&&t.iv!==''){iv=parseFloat(String(t.iv).replace('%',''));if(isNaN(iv)||iv<=0)return;}else return;
    var days=parseInt(t.days)||0;if(days<=0)return;var notl=parseFloat(t.usd_amt)||0,spot=parseFloat(t.spot)||0,strike=parseFloat(t.strike)||0,fwd=parseFloat(t.fwd_rate)||spot,ic=(t.opt_type==='CALL');
    var delta=dCD(fwd,strike,iv/100,days/365,ic);
    // Reject sub-7 delta (deep OTM with inflated smile premium)
    if(Math.abs(delta)<0.07)return;
    // Reject absurd IVs (absolute bounds)
    if(iv>60||iv<0.5)return;
    // Delta-aware vol cap for 10d bucket
    var dBucket=dDB(delta);
    if(dBucket===0||dBucket===4){
        var ti=dTB(days);
        // If broker surface exists, reject 10d prints > 3 vols from expected 10d level
        var base=dtccGetBase(pk,DT[ti]);
        if(base&&base.vols[dBucket]>0){
            if(Math.abs(iv-base.vols[dBucket])>3.5)return;
        }else{
            // No base: reject if 10d vol > 1.7x the pair's median vol so far
            var pairTrades=dAT[pk];
            if(pairTrades.length>5){
                var allIV=pairTrades.map(function(x){return x.iv;}).sort(function(a,b){return a-b;});
                var medIV=allIV[Math.floor(allIV.length/2)];
                if(iv>medIV*1.7)return;
            }
        }
    }
    if(spot>0&&(!dPM[pk]||t.time>(dPM[pk].lt||'')))dPM[pk]={spot:spot,lt:t.time};
    dAT[pk].push({time:t.time||'',type:t.opt_type||'?',strike:strike,spot:spot,fwd:fwd,days:days,iv:iv,notl:notl,expiry:t.expiry||'',ic:ic,delta:delta});}catch(e){}});}

function dBuildSurf(){dCS={};DTCC_G10.concat(DTCC_EM).forEach(function(pair){
    var trades=dAT[pair]||[],acc=[];for(var ti=0;ti<DT.length;ti++){acc[ti]=[];for(var di=0;di<5;di++)acc[ti][di]={sumIV:0,sumW:0,count:0};}
    // Pass 1: collect raw IVs per bucket for outlier detection
    var raw=[];for(var ti=0;ti<DT.length;ti++){raw[ti]=[];for(var di=0;di<5;di++)raw[ti][di]=[];}
    trades.forEach(function(t){var di=dDB(t.delta);if(di<0)return;var ti=dTB(t.days);raw[ti][di].push({iv:t.iv,w:Math.max(t.notl,1)});});
    // Pass 2: base-aware pre-filter + MAD outlier removal
    function median(arr){if(!arr.length)return null;var s=arr.slice().sort(function(a,b){return a-b;});var m=Math.floor(s.length/2);return s.length%2?s[m]:(s[m-1]+s[m])/2;}
    for(var ti=0;ti<DT.length;ti++){for(var di=0;di<5;di++){
        var ivs=raw[ti][di];if(!ivs.length)continue;
        // Base-aware pre-filter: delta-dependent threshold
        var base=dtccGetBase(pair,DT[ti]);
        if(base&&base.vols[di]){
            var expected=base.vols[di];
            // ATM: ±3 vols, 25d: ±3.5 vols, 10d: ±4 vols
            var maxDev=(di===2)?3.0:(di===1||di===3)?3.5:4.0;
            ivs=ivs.filter(function(r){return Math.abs(r.iv-expected)<=maxDev;});
        }
        if(!ivs.length)continue;
        // Single print: accept if passes base filter above (or no base)
        if(ivs.length<3){ivs.forEach(function(r){acc[ti][di].sumIV+=r.iv*r.w;acc[ti][di].sumW+=r.w;acc[ti][di].count++;});continue;}
        // Multiple prints: MAD-based filter (3x MAD, floor at 1.5 vols)
        var med=median(ivs.map(function(r){return r.iv;}));
        var mads=ivs.map(function(r){return Math.abs(r.iv-med);});var madVal=median(mads);
        var threshold=Math.max(1.5,madVal*3);
        ivs.forEach(function(r){if(Math.abs(r.iv-med)<=threshold){acc[ti][di].sumIV+=r.iv*r.w;acc[ti][di].sumW+=r.w;acc[ti][di].count++;}});
    }}
    var surf=[];for(var ti=0;ti<DT.length;ti++){var any=false;for(var di=0;di<5;di++){if(acc[ti][di].count>0){any=true;break;}}
        var base=dtccGetBase(pair,DT[ti]);
        if(base&&any){var sh=dtccShifts(base,acc[ti]);surf[ti]=dtccApply(base,sh);}
        else if(base&&!any){surf[ti]=base.vols.slice();}
        else if(!base&&any){var ae=null;if(acc[ti][2].count>0)ae=acc[ti][2].sumIV/acc[ti][2].sumW;else{var s=0,w=0;for(var di=0;di<5;di++){if(acc[ti][di].count>0){s+=acc[ti][di].sumIV;w+=acc[ti][di].sumW;}}if(w>0)ae=s/w;}
            if(ae){var gb=dtccGenBase(ae);var sh=dtccShifts(gb,acc[ti]);surf[ti]=dtccApply(gb,sh);}else surf[ti]=[null,null,null,null,null];}
        else surf[ti]=[null,null,null,null,null];}
    // Cross-tenor interp
    for(var di=0;di<5;di++){var kn=[];for(var ti=0;ti<DT.length;ti++){if(surf[ti][di]!==null)kn.push({ti:ti,v:surf[ti][di]});}if(!kn.length)continue;
        for(var ti=0;ti<DT.length;ti++){if(surf[ti][di]!==null)continue;var bl=null,ab=null;for(var k=0;k<kn.length;k++){if(kn[k].ti<ti)bl=kn[k];if(kn[k].ti>ti&&!ab)ab=kn[k];}
            if(bl&&ab){var w=(ti-bl.ti)/(ab.ti-bl.ti);surf[ti][di]=bl.v+(ab.v-bl.v)*w;}else if(bl)surf[ti][di]=bl.v;else if(ab)surf[ti][di]=ab.v;}}
    dCS[pair]=surf;});}

function dtccSnap(){dSS={};DTCC_G10.concat(DTCC_EM).forEach(function(p){if(dCS[p])dSS[p]=dCS[p].map(function(r){return r.slice();});});dST=new Date();document.getElementById('dsn').textContent='Snap '+dST.toLocaleTimeString('en-US',{hour12:false});dRA();}
function dtccClrSnap(){dSS=null;dST=null;document.getElementById('dsn').textContent='No snap';dRA();}
function dRA(){dRSum();dRSurf();dRStrat();dRTerm();dRFeed();}
function dSetSV(v){dSV=v;document.getElementById('dsv-vol').classList.toggle('active',v==='vol');document.getElementById('dsv-chg').classList.toggle('active',v==='chg');dRSurf();}
function dSetTV(v){dTV=v;['atm','rr25','fly25'].forEach(function(k){document.getElementById('dtv-'+k).classList.toggle('active',v===k);});dRTerm();}

function dRSum(){var tr=dAT[dAP]||[],sf=dCS[dAP],sn=dSS?dSS[dAP]:null;document.getElementById('ds-t').textContent=tr.length;
    var ac=0;Object.values(dAT).forEach(function(a){ac+=a.length;});document.getElementById('ds-p').textContent=ac+' | '+Object.keys(mktSurfaces).length+' marked';
    var pts=0;if(sf)sf.forEach(function(r){r.forEach(function(v){if(v!==null)pts++;});});document.getElementById('ds-pt').textContent=pts+'/'+DT.length*5;
    if(sf){for(var t=0;t<DT.length;t++){if(sf[t][2]!==null){document.getElementById('ds-af').textContent=sf[t][2].toFixed(2)+'%';document.getElementById('ds-afc').textContent=sn&&sn[t][2]!==null?((sf[t][2]-sn[t][2]>0?'+':'')+((sf[t][2]-sn[t][2]).toFixed(2))):DT[t];break;}}
        if(sf[8]&&sf[8][2]!==null){document.getElementById('ds-a1').textContent=sf[8][2].toFixed(2)+'%';document.getElementById('ds-a1c').textContent=sn&&sn[8][2]!==null?((sf[8][2]-sn[8][2]>0?'+':'')+((sf[8][2]-sn[8][2]).toFixed(2))):'1Y';}}}

function dRSurf(){var sf=dCS[dAP],sn=dSS?dSS[dAP]:null;if(!sf||sf.every(function(r){return r.every(function(v){return v===null;});})){var hm=mktSurfaces[dAP]?'Marks loaded \u2014 waiting for DTCC':'No marks for '+dAP+' \u2014 add in Market Data';document.getElementById('dsp').innerHTML='<div style="text-align:center;padding:80px 20px;color:#888">'+hm+'</div>';return;}
    var z,cs;if(dSV==='chg'&&sn){z=sf.map(function(r,t){return r.map(function(v,d){return v!==null&&sn[t][d]!==null?parseFloat((v-sn[t][d]).toFixed(3)):null;});});cs=[[0,'#1565c0'],[0.5,'#3d3d3d'],[1,'#c62828']];}
    else{z=sf.map(function(r){return r.map(function(v){return v!==null?parseFloat(v.toFixed(2)):null;});});cs=[[0,'#0d47a1'],[0.5,'#fff59d'],[1,'#b71c1c']];}
    var txt=z.map(function(r){return r.map(function(v){return v!==null?v.toFixed(2):'';});});
    Plotly.react('dsp',[{z:z,x:DDL,y:DT,type:'heatmap',colorscale:cs,text:txt,texttemplate:'%{text}',textfont:{size:10,color:'black'},hovertemplate:'%{y} %{x}<br>%{z:.2f}%<extra></extra>',connectgaps:false,colorbar:{thickness:12,len:0.85,tickfont:{size:10,color:'#aaa'}}}],{margin:{l:50,r:60,t:10,b:45},paper_bgcolor:'transparent',plot_bgcolor:'transparent',xaxis:{color:'#aaa',gridcolor:'#555',tickfont:{size:11}},yaxis:{color:'#aaa',gridcolor:'#555',tickfont:{size:11},autorange:'reversed'},font:{color:'#ccc'}},{displayModeBar:false,responsive:true});}

function dCC(v){if(v==null)return'<td class="dtcc-chg-flat">&mdash;</td>';var c=v>0.005?'dtcc-chg-up':v<-0.005?'dtcc-chg-dn':'dtcc-chg-flat';return'<td class="'+c+'">'+(v>0?'+':'')+v.toFixed(2)+'</td>';}
function dRStrat(){var sf=dCS[dAP],sn=dSS?dSS[dAP]:null,tr=dAT[dAP]||[];if(!sf){document.getElementById('dsb').innerHTML='';return;}
    var tc=Array(DT.length).fill(0);tr.forEach(function(t){tc[dTB(t.days)]++;});var rows=[];
    for(var t=0;t<DT.length;t++){var v=sf[t];if(v[2]===null)continue;var a=v[2],r25=(v[3]!=null&&v[1]!=null)?v[3]-v[1]:null,r10=(v[4]!=null&&v[0]!=null)?v[4]-v[0]:null,f25=(v[3]!=null&&v[1]!=null)?(v[3]+v[1])/2-a:null,f10=(v[4]!=null&&v[0]!=null)?(v[4]+v[0])/2-a:null;
        var da=null,dr25=null,dr10=null,df25=null,df10=null;if(sn&&sn[t]){var s=sn[t];if(s[2]!=null)da=a-s[2];if(s[3]!=null&&s[1]!=null&&r25!=null)dr25=r25-(s[3]-s[1]);if(s[4]!=null&&s[0]!=null&&r10!=null)dr10=r10-(s[4]-s[0]);if(s[3]!=null&&s[1]!=null&&f25!=null)df25=f25-((s[3]+s[1])/2-s[2]);if(s[4]!=null&&s[0]!=null&&f10!=null)df10=f10-((s[4]+s[0])/2-s[2]);}
        rows.push('<tr><td>'+DT[t]+'</td><td>'+a.toFixed(2)+'</td>'+dCC(da)+'<td>'+(r25!=null?r25.toFixed(2):'&mdash;')+'</td>'+dCC(dr25)+'<td>'+(r10!=null?r10.toFixed(2):'&mdash;')+'</td>'+dCC(dr10)+'<td>'+(f25!=null?f25.toFixed(2):'&mdash;')+'</td>'+dCC(df25)+'<td>'+(f10!=null?f10.toFixed(2):'&mdash;')+'</td>'+dCC(df10)+'<td style="color:#888">'+tc[t]+'</td></tr>');}
    document.getElementById('dsb').innerHTML=rows.join('');}

function dRTerm(){var sf=dCS[dAP],sn=dSS?dSS[dAP]:null;if(!sf){Plotly.purge('dtp');return;}var y=[],ys=[],lb='',vt=[];
    for(var t=0;t<DT.length;t++){var v=sf[t],val=null,sv=null;
        if(dTV==='atm'){val=v[2];lb='ATM (%)';if(sn&&sn[t])sv=sn[t][2];}
        else if(dTV==='rr25'){if(v[3]!=null&&v[1]!=null)val=v[3]-v[1];lb='25\u0394 RR';if(sn&&sn[t]&&sn[t][3]!=null&&sn[t][1]!=null)sv=sn[t][3]-sn[t][1];}
        else{if(v[3]!=null&&v[1]!=null&&v[2]!=null)val=(v[3]+v[1])/2-v[2];lb='25\u0394 Fly';if(sn&&sn[t]&&sn[t][3]!=null&&sn[t][1]!=null&&sn[t][2]!=null)sv=(sn[t][3]+sn[t][1])/2-sn[t][2];}
        if(val!=null){vt.push(DT[t]);y.push(parseFloat(val.toFixed(3)));if(sv!=null)ys.push(parseFloat(sv.toFixed(3)));}}
    var traces=[{x:vt,y:y,type:'scatter',mode:'lines+markers',line:{color:'#90caf9',width:2.5},marker:{size:6,color:'#90caf9'},name:'Current'}];
    if(sn&&ys.length===y.length)traces.push({x:vt,y:ys,type:'scatter',mode:'lines+markers',line:{color:'#666',width:1.5,dash:'dot'},marker:{size:4,color:'#666'},name:'Snap'});
    Plotly.react('dtp',traces,{margin:{l:45,r:15,t:5,b:30},paper_bgcolor:'transparent',plot_bgcolor:'transparent',xaxis:{color:'#aaa',gridcolor:'#555',tickfont:{size:10}},yaxis:{color:'#aaa',gridcolor:'#555',tickfont:{size:10},title:{text:lb,font:{size:10,color:'#aaa'}}},showlegend:sn?true:false,legend:{orientation:'h',y:1.12,font:{size:10,color:'#aaa'},bgcolor:'transparent'}},{displayModeBar:false,responsive:true});}

function dRFeed(){var tr=(dAT[dAP]||[]).slice().sort(function(a,b){return a.time>b.time?-1:a.time<b.time?1:0;});document.getElementById('dfc').textContent=tr.length+' prints';var list=document.getElementById('dfl');
    if(!tr.length){list.innerHTML='<div style="padding:2rem;text-align:center;color:#888">No prints</div>';return;}
    var h='';tr.slice(0,200).forEach(function(t,i){var tn=DT[dTB(t.days)],tc=t.ic?'color:#66bb6a':'color:#ef5350',vc=t.iv>=15?'color:#ffa726;font-weight:700':t.iv>=10?'color:#90caf9;font-weight:600':'color:#aaa',sc=t.notl>=50?'color:#ffa726;font-weight:700':t.notl>=20?'color:#90caf9;font-weight:600':'color:#888';
        h+='<div class="dtcc-feed-item'+(i<3?' fresh':'')+'"><span style="color:#888;font-variant-numeric:tabular-nums">'+t.time+'</span><span style="'+tc+';font-weight:700;font-size:10px">'+t.type+'</span><span style="color:#90caf9;font-weight:600">'+tn+'</span><span style="'+vc+'">'+t.iv.toFixed(1)+'</span><span style="'+sc+'">$'+t.notl+'M</span><span style="color:#888">'+Math.round(Math.abs(t.delta)*100)+'\u0394</span></div>';});
    list.innerHTML=h;}
'''

# Combine JS_ENGINE and JS_DTCC 
JS_ENGINE = JS_ENGINE + JS_DTCC

# === MAIN ===
def main():
    fp = 'fx_gamma_inputs.xlsx'
    print("\n" + "="*50 + "\n   FX OPTIONS ANALYTICS v2\n" + "="*50 + "\n")
    if not os.path.exists(fp):
        create_template(fp)
        print(f"  Created: {fp}\n  Add positions and re-run.\n")
        return
    positions = load_positions(fp)
    pairs = list(set(p['pair'] for p in positions))
    print(f"  {len(positions)} positions across {len(pairs)} pairs: {pairs}")
    out = create_dashboard(positions)
    print(f"\n  -> {out}")
    print(f"  -> Mark surfaces in Market Data tab")
    print(f"  -> Run: python serve_dashboard.py\n")

if __name__ == '__main__':
    main()
